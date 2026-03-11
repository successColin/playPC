#! /usr/bin/env python
# coding:utf-8
"""
对采集结果 Excel 按「企业名称」或「联系方式」去重，保留第一次出现的行。
用法: python deduplicate_excel.py [输入文件.xlsx] [输出文件.xlsx]
若只传一个参数则覆盖原文件；不传参数则默认处理 data_*.xlsx 中最新一个。
"""
import sys
import glob
import os
from openpyxl import load_workbook

# 用于去重的列名（优先按第一列，若不存在则按第二列）
DEDUP_COLUMN_NAMES = ('企业名称', '联系方式')


def findDedupColumnIndex(ws):
    """
    根据表头查找用于去重的列索引（1-based）。
    优先「企业名称」，其次「联系方式」。
    """
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append((col, (val or '').strip()))
    for name in DEDUP_COLUMN_NAMES:
        for col, h in headers:
            if h == name:
                return col
    return None


def deduplicateExcel(input_path, output_path=None):
    """
    读取 Excel，按去重列去重（保留首次出现的行），写回文件。
    :param input_path: 输入 xlsx 路径
    :param output_path: 输出路径，默认与输入相同（覆盖）
    """
    if output_path is None:
        output_path = input_path
    try:
        wb = load_workbook(input_path, read_only=False)
    except Exception as e:
        print(f'打开文件失败: {input_path}, 错误: {e}')
        return False
    ws = wb.active
    if ws.max_row < 2:
        print('表无数据行，无需去重')
        wb.close()
        return True
    col_idx = findDedupColumnIndex(ws)
    if col_idx is None:
        print(f'未找到去重列（表头需包含「企业名称」或「联系方式」），当前表头: {[ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column + 1, 10))]}')
        wb.close()
        return False
    seen = set()
    rows_to_remove = []  # 要删除的行号（从 2 开始）
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=col_idx).value
        key = (key or '').strip()
        if key in seen:
            rows_to_remove.append(row)
        else:
            seen.add(key)
    if not rows_to_remove:
        print('未发现重复行')
        wb.close()
        return True
    # 从后往前删行，避免行号变化
    for row in reversed(rows_to_remove):
        ws.delete_rows(row, 1)
    try:
        wb.save(output_path)
        print(f'已去除 {len(rows_to_remove)} 行重复数据，保存到: {output_path}')
    except Exception as e:
        print(f'保存失败: {e}')
        wb.close()
        return False
    wb.close()
    return True


def main():
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) >= 3 else None
    else:
        # 默认找当前目录下最新的 data_*.xlsx
        files = sorted(glob.glob(os.path.join(os.path.dirname(__file__) or '.', 'data_*.xlsx')), key=os.path.getmtime, reverse=True)
        if not files:
            print('未找到 data_*.xlsx 文件，请指定文件: python deduplicate_excel.py 输入.xlsx [输出.xlsx]')
            return
        input_path = files[0]
        output_path = None
        print(f'使用最新文件: {input_path}')
    deduplicateExcel(input_path, output_path)


if __name__ == '__main__':
    main()
