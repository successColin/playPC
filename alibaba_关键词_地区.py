#! /usr/bin/env python
# coding:utf-8

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import sys
import io
import re
from datetime import datetime
from urllib.parse import quote, urlparse
from openpyxl import Workbook
import argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── 联系信息提取相关常量 ─────────────────────────────────
# 联系方式页内用于 DOM 提取的标签关键词（与 1688 常见结构一致）
LABEL_CONTACT = '联系人'
LABEL_TEL = '电话'
LABEL_MOBILE = '手机'
LABEL_FAX = '传真'
LABEL_ADDRESS = '地址'
# 联系人正则：先匹配「联系人：xxx」，再匹配「xxx先生/女士/经理」等形式
PATTERN_CONTACT_LABEL = re.compile(r'联系人[：:]\s*([^\n]{1,20})')
PATTERN_CONTACT_SUFFIX = re.compile(r'([\u4e00-\u9fa5]{2,5})(?:先生|女士|经理|总监|总裁|老板)')
# 电话/手机/传真正则：仅保留数字、横线、空格
PATTERN_TEL = re.compile(r'电话[：:]\s*([0-9\-\u00a0\s]+)')
PATTERN_MOBILE = re.compile(r'手机[：:]\s*([0-9\-\u00a0\s]+)')
PATTERN_FAX = re.compile(r'传真[：:]\s*([0-9\-\u00a0\s]+)')
# 地址正则：到行末或遇到常见下一字段关键词为止（含「技术支持」避免抓成地址）
PATTERN_ADDRESS = re.compile(r'地址[：:]\s*([^\n]+?)(?=\s*$|邮编|传真|公司名称|邮箱|联系人|电话|手机|技术支持)', re.DOTALL)
PATTERN_ADDRESS_SIMPLE = re.compile(r'地址[：:]\s*([^\n]+)')
# 地址中需剔除的后缀（如「技术支持:旺铺管理」）
PATTERN_ADDRESS_NOISE = re.compile(r'\s*技术支持[：:][^\n]*$')
# 地址最大保留长度
MAX_ADDRESS_LEN = 200
# 搜索页公司链接若为此域名/路径则为跳转链接，需先访问取真实店铺 URL
RESOLVE_REDIRECT_HOST = 'dj.1688.com'
RESOLVE_REDIRECT_PATH = 'ci_bb'
# 1688 联系块 data-spm 容器特征（你提供的结构：电话/手机/传真/地址 每行两列 inline-block）
SPM_ANCHOR_PREFIX = 'a2615.'
# 采集结果输出文件名前缀（运行时加上时间戳，如 data_20250307_143022.xlsx）
OUTPUT_EXCEL_PREFIX = 'data'
# 搜索关键词（如 五金、模具）
SEARCH_KEYWORDS = '机械设备'
# 搜索地区筛选：省份（如 广东、浙江），空字符串表示不按省份筛选
TARGET_REGION = '广东'
# 搜索城市筛选：省份下的城市（如 广州），空字符串表示不按城市筛选
TARGET_CITY = ''
# 省份 -> 城市列表映射（按页面“从上到下、从左到右”顺序，一旦有就优先使用，不再从页面自动解析）
PROVINCE_CITY_MAP = {
    # 广东省（根据你截图中的顺序逐行展开）
    '广东': [
        '广州', '惠州', '江门',
        '深圳', '汕头', '揭阳',
        '珠海', '汕尾', '茂名',
        '潮州', '韶关', '梅州',
        '中山', '湛江', '清远',
        '东莞', '肇庆', '阳江',
        '佛山', '河源', '云浮',
    ],
}
# 部分省份城市预置映射（当页面结构变化导致自动解析失败时可直接使用）
PROVINCE_CITY_MAP = {
    # 广东省主要地级市列表（按常见顺序，可根据需要自行调整/增删）
    '广东': [
        '广州',
        '深圳',
        '珠海',
        '汕头',
        '佛山',
        '韶关',
        '湛江',
        '茂名',
        '肇庆',
        '惠州',
        '梅州',
        '汕尾',
        '河源',
        '阳江',
        '清远',
        '东莞',
        '中山',
        '潮州',
        '揭阳',
        '云浮',
    ],
}
# 每页抓取全部商家（不限制条数时用 0 或 None；正数时仅抓前 N 条，用于测试）
MAX_FETCH_PER_PAGE = 0
# Excel 表头（不含电话、传真；仅在有手机号时写入行），增加「当前城市」列
EXCEL_HEADERS = ('企业名称', '当前城市', '联系方式', '联系人', '手机', '地址')
# 滑块验证最大等待时间（秒）默认值（可通过命令行参数覆盖）
DEFAULT_CAPTCHA_WAIT_TIMEOUT = 60
# 单页全局异常最大允许次数默认值（超过后停止循环，避免极端情况下无限重试，可通过命令行参数覆盖）
DEFAULT_MAX_PAGE_ERRORS = 10
# 本次运行最多采集的商家数量（0 或负数表示不限制，可通过命令行参数覆盖）
DEFAULT_TOTAL_MAX_SHOPS = 0
# 单个店铺最少采集用时（秒），用于整体控制抓取节奏，降低触发风控风险
MIN_SECONDS_PER_SHOP = 12
# 关闭已知弹窗的 JS 片段（不关闭滑块验证框，仅关闭 baxia 等冗余遮罩）
JS_REMOVE_BAXIA_MASK = "var m=document.querySelector('.baxia-dialog-mask'); if(m) m.remove();"
JS_REMOVE_BAXIA_DIALOG = "var d=document.querySelector('.baxia-dialog'); if(d) d.remove();"
# 「亲，访问被拒绝」弹窗关键词（1688 反爬/风控提示）
TEXT_ACCESS_DENIED = '访问被拒绝'
# 访问被拒绝弹窗最大尝试关闭次数（避免死循环）
MAX_ACCESS_DENIED_CLOSE_ATTEMPTS = 3


def cleanAddress(address_str):
    """
    清洗地址字符串：去掉末尾的「技术支持:xxx」等噪音，并截断到最大长度。
    """
    if not (address_str or '').strip():
        return ''
    s = re.sub(PATTERN_ADDRESS_NOISE, '', address_str).strip()
    return s[:MAX_ADDRESS_LEN] if s else ''


def loadRuntimeConfig():
    """
    通过命令行参数加载运行配置（验证码超时时间 / 页面异常上限 / 本次最多采集数量）。
    若未传入对应参数，则使用默认常量值，方便根据当天风控情况灵活调整。
    """
    global CAPTCHA_WAIT_TIMEOUT, MAX_PAGE_ERRORS, TOTAL_MAX_SHOPS

    parser = argparse.ArgumentParser(
        description='1688 商家联系方式采集脚本运行参数'
    )
    parser.add_argument(
        '--captcha-timeout',
        type=int,
        default=DEFAULT_CAPTCHA_WAIT_TIMEOUT,
        help='滑块验证码最长等待时间（秒），默认 60 秒',
    )
    parser.add_argument(
        '--max-page-errors',
        type=int,
        default=DEFAULT_MAX_PAGE_ERRORS,
        help='分页循环中允许的最大页面级异常次数，默认 10 次',
    )
    parser.add_argument(
        '--max-shops',
        type=int,
        default=DEFAULT_TOTAL_MAX_SHOPS,
        help='本次运行最多采集的商家数量，0 或负数表示不限制，适合测试时只抓前 N 家',
    )

    # 使用 parse_known_args，避免未来需要在 sys.argv 中加入其他参数时出错
    args, _ = parser.parse_known_args()

    # 根据命令行参数更新全局运行时配置
    CAPTCHA_WAIT_TIMEOUT = args.captcha_timeout if args.captcha_timeout > 0 else DEFAULT_CAPTCHA_WAIT_TIMEOUT
    MAX_PAGE_ERRORS = args.max_page_errors if args.max_page_errors > 0 else DEFAULT_MAX_PAGE_ERRORS
    TOTAL_MAX_SHOPS = args.max_shops if args.max_shops >= 0 else DEFAULT_TOTAL_MAX_SHOPS


def getShopOrigin(shop_url):
    """
    从任意店铺 URL 解析出「协议 + 域名」，用于拼接 /page/contactinfo.htm。
    避免带路径的 URL（如 .../page/main.htm）被错误拼接成 .../page/main.htm/page/contactinfo.htm。
    """
    if not (shop_url or '').strip():
        return ''
    try:
        parsed = urlparse(shop_url.strip())
        if parsed.scheme and parsed.netloc:
            return f'{parsed.scheme}://{parsed.netloc}'
    except Exception:
        pass
    return shop_url.rstrip('/').split('/page/')[0].split('?')[0] or shop_url


def closeKnownPopups(driver):
    """
    关闭已知的冗余弹窗（如 baxia 遮罩/对话框），不关闭滑块验证框，避免多框叠加。
    在打开新窗口或加载联系页后调用，减少「弹出多个框」的干扰。
    """
    try:
        driver.execute_script(JS_REMOVE_BAXIA_MASK)
        driver.execute_script(JS_REMOVE_BAXIA_DIALOG)
    except Exception:
        pass


def closeAccessDeniedPopup(driver):
    """
    检测并尝试关闭「亲，访问被拒绝」弹窗（1688 风控提示）。
    若页面存在该文案，则尝试点击关闭按钮或移除弹窗 DOM，便于继续抓取。
    返回 True 表示曾检测到并已尝试关闭，False 表示未发现该弹窗。
    """
    try:
        body_text = (driver.find_element(By.TAG_NAME, 'body').text or '')
        if TEXT_ACCESS_DENIED not in body_text:
            return False
        # 尝试多种方式关闭：先找关闭按钮点击，再尝试移除弹窗容器
        closed = False
        for _ in range(MAX_ACCESS_DENIED_CLOSE_ATTEMPTS):
            try:
                # 方式1：通过包含「访问被拒绝」的文案找到弹窗，再找其内的关闭按钮（常见为 × 或 class 含 close）
                deny_els = driver.find_elements(
                    By.XPATH,
                    "//*[contains(text(),'" + TEXT_ACCESS_DENIED + "')]"
                )
                for el in deny_els:
                    try:
                        # 在弹窗容器内找关闭按钮：同一父级下的 button 或 a 或 span（文案为 × 或 关闭）
                        parent = el.find_element(
                            By.XPATH,
                            "./ancestor::*[contains(@class,'dialog') or contains(@class,'modal') or contains(@class,'popup')][1]"
                        )
                        close_btns = parent.find_elements(
                            By.XPATH,
                            ".//*[contains(@class,'close') or text()='×' or text()='关闭' or contains(text(),'×')]"
                        )
                        if close_btns:
                            driver.execute_script("arguments[0].click();", close_btns[0])
                            closed = True
                            time.sleep(1)
                            break
                    except Exception:
                        pass
                if closed:
                    break
                # 方式2：用 JS 移除包含「访问被拒绝」的弹窗及其遮罩（通过常见 class 或标签）
                script = """
                var text = '""" + TEXT_ACCESS_DENIED + """';
                var all = document.querySelectorAll('div, section');
                for (var i = all.length - 1; i >= 0; i--) {
                    var el = all[i];
                    if (el.innerText && el.innerText.indexOf(text) !== -1) {
                        var p = el.closest('.dialog') || el.closest('.modal') || el.closest('[class*="dialog"]') || el.closest('[class*="modal"]') || el.parentElement;
                        if (p) { p.remove(); return true; }
                    }
                }
                return false;
                """
                removed = driver.execute_script(script)
                if removed:
                    closed = True
                    break
            except Exception:
                pass
            time.sleep(0.5)
        if closed:
            time.sleep(1)
        return closed
    except Exception:
        return False


def resolveShopUrl(driver, href_value):
    """
    若链接为 1688 跳转链接（dj.1688.com/ci_bb），先访问一次取重定向后的真实店铺 URL，
    再用于拼接 /page/contactinfo.htm，避免联系方式页打开错误。
    返回可用于拼接联系页的店铺 base URL（可能带路径，后续由 getShopOrigin 取纯域名）。
    """
    if not (href_value or '').strip():
        return href_value or ''
    href = href_value.strip().rstrip('/')
    # 判断是否为跳转链接（非真实店铺页）
    is_redirect = (RESOLVE_REDIRECT_HOST in href) or (RESOLVE_REDIRECT_PATH in href)
    if not is_redirect:
        return href
    try:
        driver.get(href)
        time.sleep(random.uniform(2, 4))
        real_url = driver.current_url or href
        # 若跳转后仍是异常 URL，保留原 href 避免死链
        if real_url and '.1688.com' in real_url and 'ci_bb' not in real_url:
            return real_url.rstrip('/')
        return href
    except Exception:
        return href


def extractContactByDom(driver):
    """
    从当前页面 DOM 中提取联系人、电话、手机、传真、地址。
    优先匹配 1688 联系块结构：带 data-spm-anchor-id 的容器内，每行「标签：」+ 值（两列 inline-block）。
    返回 (member_name, tel, mobile, fax, address)，未找到的项为 ''。
    """
    member_name = ''
    tel = ''
    mobile = ''
    fax = ''
    address = ''
    try:
        # 方法0：1688 联系块结构——父级 div[data-spm-anchor-id] 内多行，每行两列（电话：/手机：/传真：/地址： + 值）
        try:
            containers = driver.find_elements(
                By.XPATH,
                "//div[contains(@data-spm-anchor-id,'" + SPM_ANCHOR_PREFIX + "') and .//div[contains(text(),'电话：')]]"
            )
            for container in containers:
                # 每行是容器的直接子 div，行内两个子 div 分别为标签、值
                rows = container.find_elements(By.XPATH, "./div")
                for row in rows:
                    parts = row.find_elements(By.XPATH, "./div")
                    if len(parts) < 2:
                        continue
                    label_text = (parts[0].text or '').strip()
                    value_el = parts[1]
                    value_text = (value_el.text or '').strip()
                    # 地址可能被省略显示，完整内容在 title 中
                    if LABEL_ADDRESS in label_text or label_text == '地址：' or label_text == '地址':
                        title_addr = value_el.get_attribute('title')
                        if title_addr and title_addr.strip():
                            value_text = title_addr.strip()
                        if value_text and not address:
                            address = value_text[:MAX_ADDRESS_LEN]
                    elif LABEL_TEL in label_text or label_text == '电话：' or label_text == '电话':
                        if value_text and not tel:
                            tel = re.sub(r'\s+', ' ', re.sub(r'[^\d\-\s]', '', value_text)).strip()[:50]
                    elif LABEL_MOBILE in label_text or label_text == '手机：' or label_text == '手机':
                        if value_text and not mobile:
                            mobile = re.sub(r'\s+', ' ', re.sub(r'[^\d\-\s]', '', value_text)).strip()[:50]
                    elif LABEL_FAX in label_text or label_text == '传真：' or label_text == '传真':
                        if value_text and not fax:
                            fax = re.sub(r'\s+', ' ', re.sub(r'[^\d\-\s]', '', value_text)).strip()[:50]
                if tel or mobile or fax or address:
                    break
        except Exception:
            pass

        # 方法1：dl > dt + dd 成对（常见于 1688 联系信息块）
        if not tel and not mobile and not fax and not address:
            dts = driver.find_elements(By.TAG_NAME, 'dt')
            for dt in dts:
                label = (dt.text or '').strip()
                try:
                    dd = dt.find_element(By.XPATH, './following-sibling::dd[1]')
                    value = (dd.text or '').strip()
                except Exception:
                    value = ''
                if not value:
                    continue
                if LABEL_CONTACT in label or label == '联系人':
                    member_name = value[:30]
                elif LABEL_TEL in label or label == '电话':
                    digits = re.sub(r'\s+', ' ', re.sub(r'[^\d\-]', '', value)).strip()
                    if len(re.sub(r'[^\d]', '', digits)) >= 5:
                        tel = digits[:50]
                elif LABEL_MOBILE in label or label == '手机':
                    digits = re.sub(r'\s+', ' ', re.sub(r'[^\d\-]', '', value)).strip()
                    if len(re.sub(r'[^\d]', '', digits)) >= 5:
                        mobile = digits[:50]
                elif LABEL_FAX in label or label == '传真':
                    digits = re.sub(r'\s+', ' ', re.sub(r'[^\d\-]', '', value)).strip()
                    if digits:
                        fax = digits[:50]
                elif LABEL_ADDRESS in label or label == '地址':
                    address = value[:MAX_ADDRESS_LEN]
        # 方法2：通过 class 含关键词的块（tel/phone/contact/address/member）
        if not member_name or not tel or not address:
            for xpath_label, key in [
                ("//*[contains(@class,'member') or contains(@class,'contact-name')]", 'member'),
                ("//*[contains(@class,'tel') or contains(@class,'phone')]", 'tel'),
                ("//*[contains(@class,'address') or contains(@class,'addr')]", 'address'),
            ]:
                try:
                    els = driver.find_elements(By.XPATH, xpath_label)
                    for el in els:
                        t = (el.text or '').strip()
                        if not t or len(t) > 300:
                            continue
                        if key == 'member' and not member_name and re.match(r'^[\u4e00-\u9fa5a-zA-Z\s]{2,20}$', t):
                            member_name = t[:30]
                            break
                        if key == 'tel' and not tel and re.search(r'\d{5,}', t):
                            tel = re.sub(r'\s+', '-', re.sub(r'[^\d\-]', '', t))[:50] or t[:50]
                            break
                        if key == 'address' and not address:
                            address = t[:MAX_ADDRESS_LEN]
                            break
                except Exception:
                    pass
        # 方法3：整段文本中带「联系人：」「电话：」「手机：」「传真：」「地址：」的块
        if not member_name or not tel or not mobile or not fax or not address:
            full_body = (driver.find_element(By.TAG_NAME, 'body').text or '')
            if not member_name:
                rm = PATTERN_CONTACT_LABEL.search(full_body)
                if rm:
                    member_name = rm.group(1).strip()[:30]
            if not tel:
                rm = PATTERN_TEL.search(full_body)
                if rm:
                    tel = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', rm.group(1))).strip()[:50]
            if not mobile:
                rm = PATTERN_MOBILE.search(full_body)
                if rm:
                    mobile = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', rm.group(1))).strip()[:50]
            if not fax:
                rm = PATTERN_FAX.search(full_body)
                if rm:
                    fax = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', rm.group(1))).strip()[:50]
            if not address:
                rm = PATTERN_ADDRESS.search(full_body)
                if rm:
                    address = rm.group(1).strip()[:MAX_ADDRESS_LEN]
                if not address:
                    rm = PATTERN_ADDRESS_SIMPLE.search(full_body)
                    if rm:
                        address = rm.group(1).strip()[:MAX_ADDRESS_LEN]
    except Exception:
        pass
    return (member_name or '', tel or '', mobile or '', fax or '', address or '')


def extractContactByRegex(page_text):
    """
    从页面纯文本中用正则提取联系人、电话、手机、传真、地址。用于 DOM 取不到时的兜底。
    返回 (member_name, tel, mobile, fax, address)，未找到的项为 ''。
    """
    member_name = ''
    tel = ''
    mobile = ''
    fax = ''
    address = ''
    if not (page_text or '').strip():
        return ('', '', '', '', '')
    text = page_text
    # 联系人：优先「联系人：xxx」
    m = PATTERN_CONTACT_LABEL.search(text)
    if m:
        member_name = m.group(1).strip()
        member_name = re.sub(r'[\s\d].*', '', member_name)[:20]
    if not member_name:
        m = PATTERN_CONTACT_SUFFIX.search(text)
        if m:
            # 正则只有一组捕获（中文名），(?:先生|女士|...) 为非捕获组，无 group(2)
            member_name = m.group(1).strip()[:30]
    # 电话
    m = PATTERN_TEL.search(text)
    if m:
        tel = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]
    # 手机
    m = PATTERN_MOBILE.search(text)
    if m:
        mobile = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]
    # 传真
    m = PATTERN_FAX.search(text)
    if m:
        fax = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]
    # 地址
    m = PATTERN_ADDRESS.search(text)
    if m:
        address = m.group(1).strip()
    if not address:
        m = PATTERN_ADDRESS_SIMPLE.search(text)
        if m:
            address = m.group(1).strip()
    address = (address or '')[:MAX_ADDRESS_LEN]
    return (member_name or '', tel or '', mobile or '', fax or '', address or '')


# 创建浏览器驱动：优先 Chrome（兼容性与速度更好），不可用时使用 Firefox
def createBrowserDriver():
    """优先创建 Chrome 驱动，失败则使用 Firefox。"""
    try:
        driver = webdriver.Chrome()
        print('使用浏览器: Chrome')
        return driver
    except Exception as e1:
        try:
            driver = webdriver.Firefox()
            print('使用浏览器: Firefox (Chrome 不可用:', str(e1)[:50], ')')
            return driver
        except Exception as e2:
            raise RuntimeError(f'Chrome 与 Firefox 均不可用。Chrome: {e1}; Firefox: {e2}') from e2


def buildOutputFileName():
    """
    根据当前日期时间、搜索关键词和地区构造导出 Excel 文件名。
    文件命名格式：YYYYMMDD_HHMMSS_搜索内容_地区.excel，例如：20260309_153045_五金_广东广州.excel。
    若未设置地区，则地区部分使用「全国」。
    """
    try:
        # 构造时间部分：精确到秒，方便区分多次运行
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 构造地区部分：省 + 市，如「广东广州」，未配置时使用「全国」
        region_parts = []
        if TARGET_REGION:
            region_parts.append(TARGET_REGION)
            if TARGET_CITY:
                region_parts.append(TARGET_CITY)
        region_desc = ''.join(region_parts) if region_parts else '全国'
        # 构造搜索内容部分：去掉中间空格，避免文件名中出现多余空格
        keywords_part = (SEARCH_KEYWORDS or '未命名').replace(' ', '')
        # 最终文件名：日期时间_搜索内容_地区.xlsx
        return f"{now_str}_{keywords_part}_{region_desc}.xlsx"
    except Exception:
        # 兜底：如果拼接出错，退回原来的前缀 + 时间命名方式
        return f"{OUTPUT_EXCEL_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def waitCaptchaResolved(driver):
    """
    使用 WebDriverWait 轮询等待滑块验证码消失。
    返回 (resolved, waited_seconds)：
    - resolved 为 True 表示在超时时间内验证码已消失；
    - resolved 为 False 表示在 CAPTCHA_WAIT_TIMEOUT 秒内验证码仍存在。
    """
    start_ts = time.time()
    # 第一次检查当前页面是否已经出现验证码
    try:
        body_text = driver.find_element(By.TAG_NAME, 'body').text or ''
    except Exception:
        body_text = ''
    has_captcha = (
        'slide to verify' in body_text.lower()
        or '滑动验证' in body_text
        or '拖动' in body_text
    )
    if not has_captcha:
        # 当前页面根本没有验证码，直接认为已“通过”
        return True, 0

    print('  ⚠ 出现滑块验证码！请在浏览器中手动拖动滑块完成验证，脚本自动等待...')
    # 先尝试关闭其他已知弹窗，避免遮挡验证码区域
    try:
        closeKnownPopups(driver)
    except Exception:
        pass

    # 使用独立的 WebDriverWait，在 CAPTCHA_WAIT_TIMEOUT 内轮询等待验证码消失
    wait_captcha = WebDriverWait(driver, CAPTCHA_WAIT_TIMEOUT, poll_frequency=3.0)

    def _captcha_gone(d):
        """内部轮询函数：当页面上不再包含验证码提示文案时返回 True。"""
        try:
            txt = d.find_element(By.TAG_NAME, 'body').text or ''
        except Exception:
            return False
        has_flag = (
            'slide to verify' in txt.lower()
            or '滑动验证' in txt
            or '拖动' in txt
        )
        if has_flag:
            # 每次轮询时顺带尝试关闭除验证码外的已知弹窗
            try:
                closeKnownPopups(d)
            except Exception:
                pass
            return False
        return True

    try:
        wait_captcha.until(_captcha_gone)
        waited = int(time.time() - start_ts)
        return True, waited
    except Exception:
        waited = int(time.time() - start_ts)
        return False, waited


def scrollToBottom(driver):
    """
    将当前页面平滑滚动到底部。
    使用 window.scrollTo + document.body.scrollHeight，避免依赖固定像素高度（如 30000）导致某些分辨率下无法完全滚动到底部。
    """
    try:
        driver.execute_script(
            "window.scrollTo({top: document.body.scrollHeight, behavior: 'auto'});"
        )
    except Exception:
        # 若浏览器不支持上述参数形式，则降级为最简单的 scrollTo 调用
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        except Exception:
            pass


def getCityListByProvince(driver, province_name):
    """
    根据指定省份，自动解析 1688 搜索页「所在地区」筛选中的城市列表顺序。

    设计目标：
    1. 省份固定时，只要修改 TARGET_REGION，脚本即可按「从上到下，从左到右」的视觉顺序依次抓取该省下所有城市。
    2. 通过元素在页面中的位置信息（top、left）进行排序，尽量贴合实际展示顺序。

    参数说明：
    - driver: 已登录状态下的 Selenium WebDriver 实例。
    - province_name: 省份名称，如「广东」「浙江」等。

    返回值：
    - 返回城市名称字符串列表，例如 ["广州", "惠州", "江门", "深圳", ...]。
      若解析失败，返回空列表，调用方需自行兜底（例如退化为整省不分市抓取）。
    """
    city_list = []
    # 第 0 步：若该省在预置映射中，优先直接返回预置城市列表，避免因 DOM 结构变更导致解析失败
    try:
        preset_cities = PROVINCE_CITY_MAP.get(province_name)
        if preset_cities:
            print(f'省份「{province_name}」使用预置城市列表: {preset_cities}')
            return list(preset_cities)
    except Exception:
        pass
    try:
        # 构造仅指定省份、不指定城市的搜索 URL
        base_url = (
            'https://s.1688.com/company/company_search.htm?'
            'keywords=' + quote(SEARCH_KEYWORDS, encoding='gbk', safe='')
            + '&n=y&spm=a260k.635.1998096057.d1'
            + '&province=' + quote(province_name, encoding='gbk', safe='')
        )
        driver.get(base_url)
        # 等待页面关键筛选区域加载完成（所在地区筛选）
        try:
            temp_wait = WebDriverWait(driver, 20)
            temp_wait.until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "div.sm-widget-address, div.sm-widget-region, div.address-widget")
                )
            )
        except Exception:
            # 即使等待失败，也继续尝试查找元素，避免因为 DOM 结构差异导致直接中断
            pass

        # 有些情况下城市列表需要鼠标悬停「广东」省份才能展开，
        # 这里尝试点击一次对应省份的链接，促使右侧城市列表弹出。
        try:
            province_elems = driver.find_elements(
                By.XPATH,
                "//*[text()='" + province_name + "']"
            )
            for prov in province_elems:
                try:
                    driver.execute_script("arguments[0].click();", prov)
                    time.sleep(1)
                    break
                except Exception:
                    continue
        except Exception:
            pass

        # 在地区筛选区域内查找所有城市链接元素。
        # 第 1 步：优先在「所在地区」组件常见容器内查找；
        # 第 2 步：若未找到，则退化为全局查找带 &city= 参数的公司搜索链接，进一步放宽条件。
        candidate_city_elements = []
        try:
            containers = driver.find_elements(
                By.XPATH,
                "//div[contains(@class,'sm-widget-address') or contains(@class,'sm-widget-region') or contains(@class,'address-widget')]"
            )
            for container in containers:
                links = container.find_elements(By.TAG_NAME, "a")
                for link in links:
                    text_val = (link.text or '').strip()
                    # 城市名通常为 1~6 个中文字符，排除「不限」「全部」「全国」等无效项
                    if not text_val:
                        continue
                    if text_val in ('不限', '全部', '全国'):
                        continue
                    if not re.match(r'^[\u4e00-\u9fa5]{1,6}$', text_val):
                        continue
                    try:
                        rect = driver.execute_script(
                            "var r = arguments[0].getBoundingClientRect(); return [r.top, r.left];",
                            link,
                        )
                        top_pos = float(rect[0]) if rect and len(rect) >= 1 else 0.0
                        left_pos = float(rect[1]) if rect and len(rect) >= 2 else 0.0
                    except Exception:
                        top_pos = 0.0
                        left_pos = 0.0
                    candidate_city_elements.append((top_pos, left_pos, text_val))
        except Exception:
            candidate_city_elements = []

        # 若在常见容器中未找到城市元素，则进一步做一次全局兜底：
        # 1. href 中包含 company_search.htm 与 &city=；
        # 2. 文本为 1~6 个中文字符，排除「不限」「全部」「全国」等。
        if not candidate_city_elements:
            try:
                all_links = driver.find_elements(By.XPATH, "//a[@href and contains(@href,'company_search.htm') and contains(@href,'city=')]")
                for link in all_links:
                    text_val = (link.text or '').strip()
                    if not text_val:
                        continue
                    if text_val in ('不限', '全部', '全国'):
                        continue
                    if not re.match(r'^[\u4e00-\u9fa5]{1,6}$', text_val):
                        continue
                    try:
                        rect = driver.execute_script(
                            "var r = arguments[0].getBoundingClientRect(); return [r.top, r.left];",
                            link,
                        )
                        top_pos = float(rect[0]) if rect and len(rect) >= 1 else 0.0
                        left_pos = float(rect[1]) if rect and len(rect) >= 2 else 0.0
                    except Exception:
                        top_pos = 0.0
                        left_pos = 0.0
                    candidate_city_elements.append((top_pos, left_pos, text_val))
            except Exception:
                candidate_city_elements = []

        if not candidate_city_elements:
            # 打印一条简单日志，方便你在终端看到当前页面结构是否被成功识别
            print(f'未在省份「{province_name}」页面上找到城市链接元素，城市列表为空')
            return []

        # 排序方式：先按 top（从小到大），再按 left（从小到大），
        # 对应视觉上的「从上到下，从左到右」顺序。
        candidate_city_elements.sort(key=lambda item: (round(item[0], 1), round(item[1], 1)))

        # 去重并保持顺序
        seen_names = set()
        for _, _, name in candidate_city_elements:
            if name not in seen_names:
                seen_names.add(name)
                city_list.append(name)
    except Exception as e:
        # 这里仅打印提示，不抛出异常，避免影响后续整体采集流程
        print(f'自动解析省份「{province_name}」城市列表失败: {e}')

    return city_list


# 先加载运行参数，再创建浏览器驱动
loadRuntimeConfig()
driver = createBrowserDriver()
wait = WebDriverWait(driver, 120)

# 公司搜索页基础 URL（关键词用 GBK 编码，与 1688 一致）
_BASE_SEARCH_URL = (
    'https://s.1688.com/company/company_search.htm?'
    'keywords=' + quote(SEARCH_KEYWORDS, encoding='gbk', safe='') + '&n=y&spm=a260k.635.1998096057.d1'
)


def buildSearchUrl(province_name, city_name):
    """
    根据省份与城市构造公司搜索页 URL。

    说明：
    - 省份固定时，只要切换 city_name，即可依次抓取该省旗下不同城市。
    - 当 city_name 为空字符串时，表示按整省（不区分城市）进行搜索。
    """
    try:
        base_url = _BASE_SEARCH_URL
        if province_name:
            base_url = base_url + '&province=' + quote(province_name, encoding='gbk', safe='')
            if city_name:
                base_url = base_url + '&city=' + quote(city_name, encoding='gbk', safe='')
        return base_url
    except Exception:
        # 若拼接过程中出错，退回到最基础的关键词搜索 URL
        return _BASE_SEARCH_URL

# ── 扫码登录 ────────────────────────────────────────────
# 打开淘宝登录页（默认显示扫码界面）
driver.get('https://login.taobao.com/member/login.jhtml')

print('=' * 50)
print('请打开手机淘宝 App，扫描浏览器中的二维码完成登录')
print('等待扫码中（最多等待 120 秒）...')
print('=' * 50)

# 等待登录成功：检测页面跳转离开登录域名
try:
    wait.until(EC.url_contains('taobao.com/'))
    # 确保不再停留在 login 页面
    wait.until_not(EC.url_contains('login.taobao.com'))
    print('登录成功！')
except Exception:
    # 超时后检查当前URL，如果已经跳走也算成功
    if 'login.taobao.com' not in driver.current_url:
        print('登录成功！')
    else:
        print('登录超时，请重新运行脚本并及时扫码')
        driver.quit()
        sys.exit(1)

# ── 根据省份自动生成城市列表 ─────────────────────────────
if TARGET_REGION:
    if TARGET_CITY:
        # 若手动指定了城市，则只抓取该城市
        city_list = [TARGET_CITY]
    else:
        # 未指定城市时，自动解析该省下所有城市（从上到下、从左到右）
        city_list = getCityListByProvince(driver, TARGET_REGION)
        if not city_list:
            print(f'未能自动解析省份「{TARGET_REGION}」的城市列表，将按整省抓取一次')
            city_list = ['']
        else:
            # 打印当前省份下解析到的城市数量，便于确认抓取范围（立即刷新输出）
            effective_cities = [c for c in city_list if c]
            print(f'省份「{TARGET_REGION}」共解析到 {len(effective_cities)} 个城市: {city_list}', flush=True)
else:
    # 未指定省份时，仅按全国维度抓取一次
    city_list = ['']

print(f'本次将按以下城市顺序抓取（空字符串表示整省/全国一次）: {city_list}')

# 生成带时间戳的输出文件名，避免多次运行覆盖
OUTPUT_EXCEL = buildOutputFileName()
# 新建 Excel 工作簿并写入表头
wb = Workbook()
ws = wb.active
ws.title = '采集数据'
for col, header in enumerate(EXCEL_HEADERS, start=1):
    ws.cell(row=1, column=col, value=header)
excel_row = 2
# 已采集店铺集合（用 shop_origin 去重，避免同一店铺多页/多链接重复写入）
seen_shops = set()
# 本次运行已成功采集的商家数量（跨所有城市累计）
total_shops_collected = 0

# ── 按城市外层循环，内层保留原有分页抓取逻辑 ─────────────
for current_city in city_list:
    # 更新全局当前城市，便于日志输出与文件命名中使用
    TARGET_CITY = current_city or ''
    search_url = buildSearchUrl(TARGET_REGION, TARGET_CITY)

    # 打开对应城市的搜索首页，主标签页始终保留此搜索结果页；联系页在新建标签中打开，抓完即关
    driver.get(search_url)
    if TARGET_REGION:
        region_desc = TARGET_REGION + (' - ' + TARGET_CITY if TARGET_CITY else '')
        print(f'已设置地区筛选: {region_desc}')
    else:
        print('未设置地区筛选（全国范围）')

    # 主标签句柄：用于采集时在新标签打开联系页，抓完后关闭新标签并切回主标签
    main_window = driver.current_window_handle
    # 当前城市下的页面异常计数与控制标记
    page_error_count = 0
    stop_collecting = False
    # 循环抓取每一页，直到没有下一页或达到采集上限
    page_num = 1

    while True:
        if stop_collecting:
            break
        try:
            # 每页开始时先关闭已知弹窗，并尝试关闭「亲，访问被拒绝」弹窗以便继续抓取
            closeKnownPopups(driver)
            if closeAccessDeniedPopup(driver):
                print('检测到「访问被拒绝」弹窗，已尝试关闭；若列表仍为空请手动关闭弹窗或稍后重试')
            # 使用显式等待，确保当前页公司列表元素加载完成，减少直接 find_elements 导致空列表的风险
            try:
                wait.until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "a.company-name")
                    )
                )
            except Exception:
                pass
            # 获取当前页企业名称列表
            title = driver.find_elements(By.CSS_SELECTOR, "a.company-name")
            # 打印当前页商家数量；若为 0 且页面有「访问被拒绝」提示，给出说明
            if len(title) == 0:
                try:
                    body_text = (driver.find_element(By.TAG_NAME, 'body').text or '')
                    if TEXT_ACCESS_DENIED in body_text:
                        print(f'第{page_num}页列表为空，且存在「访问被拒绝」提示，请手动关闭弹窗或扫码验证后重试')
                except Exception:
                    pass
            print(f'当前城市 {TARGET_CITY or "整省/全国"} 第{page_num}页找到 {len(title)} 个商家')
            # 尝试解析分页控件中的总页数：如果当前页数已经等于或超过总页数，则结束当前城市的采集
            try:
                pager_text = ''
                # 常见分页容器 class 名称做一个并集匹配，尽量兼容不同页面结构
                pager_elements = driver.find_elements(
                    By.CSS_SELECTOR,
                    "div.fui-pager, div.sm-pagination, div[class*='pagination'], span.page-count"
                )
                for el in pager_elements:
                    text_val = (el.text or '').strip()
                    if text_val:
                        pager_text = text_val
                        break
                total_pages = 0
                if pager_text:
                    # 兼容类似「1/52」或「第1页/共52页」等样式
                    m = re.search(r'/\s*(\d+)', pager_text)
                    if not m:
                        m = re.search(r'共\s*(\d+)\s*页', pager_text)
                    if m:
                        total_pages = int(m.group(1))
                # 如能解析出总页数，则在日志中打印出来，方便观察当前城市一共多少页
                if total_pages > 0:
                    print(
                        f'城市 {TARGET_CITY or "整省/全国"} 当前搜索结果共 {total_pages} 页',
                        flush=True
                    )
                if total_pages > 0 and page_num >= total_pages:
                    # 当当前页码已经等于分页控件上的总页数时，认为已经到最后一页，直接结束当前城市循环
                    print(
                        f'城市 {TARGET_CITY or "整省/全国"} 当前第{page_num}页已达到分页控件显示的末页（共 {total_pages} 页），结束该城市采集'
                    )
                    break
            except Exception:
                # 分页信息解析失败时不影响正常翻页逻辑
                pass
            # 先收集当前页所有商家信息（包含元素本身和其在视口中的垂直位置），避免跳转后引用失效
            page_data = []
            for i in range(len(title)):
                title_el = title[i]
                title_value = title_el.get_attribute('title') or title_el.text
                href_value = title_el.get_attribute('href')
                if not href_value:
                    continue
                # 使用 getBoundingClientRect().top 记录当前公司名称链接在页面中的垂直位置，
                # 后续按该位置排序，保证抓取顺序与页面由上到下的视觉顺序一致。
                try:
                    top_pos = driver.execute_script(
                        "return arguments[0].getBoundingClientRect().top;", title_el
                    )
                except Exception:
                    top_pos = 0
                # 结构： (垂直位置, 公司链接元素, 公司标题文本, 公司链接)
                page_data.append((top_pos, title_el, title_value, href_value))

            # 按垂直位置从小到大排序，确保抓取顺序与页面视觉顺序一致
            page_data.sort(key=lambda x: x[0])

            # 当前页待抓取列表（若 MAX_FETCH_PER_PAGE>0 则只抓前 N 条，否则抓本页全部）
            ordered_data = page_data[:MAX_FETCH_PER_PAGE] if MAX_FETCH_PER_PAGE else page_data
            for _, title_el, title_value, href_value in ordered_data:
                # 在打开新标签采集前，将当前这条商家记录滚动到搜索结果列表的可视区域中间，方便人工观察采集进度
                try:
                    if title_el:
                        # 直接对当前这条“公司名称”链接元素执行滚动，block='center' 让其尽量出现在视口中间
                        driver.execute_script(
                            "arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});",
                            title_el
                        )
                except Exception:
                    # 若元素已失效或滚动异常，不影响后续采集流程，直接忽略
                    pass
                # 记录当前店铺采集开始时间，用于控制单条最少耗时，防止抓取过快
                shop_start_ts = time.time()
                # print(f'正在采集: {title_value}')
                try:
                    href = (href_value or '').strip()
                    is_redirect = (RESOLVE_REDIRECT_HOST in href) or (RESOLVE_REDIRECT_PATH in href)

                    if not is_redirect:
                        # 非跳转链接：直接在新标签打开联系页，避免先打开店铺首页再跳转，减少弹窗次数
                        shop_origin = getShopOrigin(href)
                        dedup_key = (shop_origin or href) or title_value
                        if dedup_key in seen_shops:
                            print(f'  跳过重复店铺（已采集）: {title_value}')
                            continue
                        if not shop_origin:
                            print(f'  无法解析店铺域名，跳过')
                            continue
                        contact_url = shop_origin.rstrip('/') + '/page/contactinfo.htm'
                        driver.execute_script("window.open(arguments[0], '_blank');", contact_url)
                        driver.switch_to.window(driver.window_handles[-1])
                        seen_shops.add(dedup_key)
                        closeKnownPopups(driver)
                    else:
                        # 跳转链接：必须先在新标签打开一次以解析真实 URL，再打开联系页
                        driver.execute_script("window.open(arguments[0], '_blank');", href)
                        driver.switch_to.window(driver.window_handles[-1])
                        closeKnownPopups(driver)
                        # 新标签已打开跳转链接并完成重定向，直接从当前 URL 解析店铺域名，避免再次 get 触发弹窗
                        shop_origin = getShopOrigin((driver.current_url or '').strip())
                        dedup_key = (shop_origin or href) or title_value
                        if dedup_key in seen_shops:
                            print(f'  跳过重复店铺（已采集）: {title_value}')
                            driver.close()
                            driver.switch_to.window(main_window)
                            continue
                        if not shop_origin:
                            print(f'  无法解析店铺域名，跳过')
                            driver.close()
                            driver.switch_to.window(main_window)
                            continue
                        contact_url = shop_origin.rstrip('/') + '/page/contactinfo.htm'
                        driver.get(contact_url)
                        closeKnownPopups(driver)
                        seen_shops.add(dedup_key)

                    # 若未真正进入联系页（被重定向到首页/验证等），用当前页域名再试一次联系页
                    current_url = (driver.current_url or '').strip()
                    if 'contactinfo' not in current_url and '.1688.com' in current_url:
                        retry_origin = getShopOrigin(current_url)
                        if retry_origin:
                            retry_contact = retry_origin.rstrip('/') + '/page/contactinfo.htm'
                            driver.get(retry_contact)
                        closeKnownPopups(driver)

                    # 仍未在联系页时提示，便于排查
                    if 'contactinfo' not in (driver.current_url or ''):
                        print(f'  ⚠ 未进入联系方式页，当前: {(driver.current_url or "")[:80]}')

                    # 使用显式等待确保联系方式页主体 <body> 已渲染，再读取文本内容
                    try:
                        wait.until(
                            EC.presence_of_element_located((By.TAG_NAME, 'body'))
                        )
                    except Exception:
                        pass
                    # 获取渲染后的页面文本
                    page_text = driver.find_element(By.TAG_NAME, 'body').text

                    # 检查是否出现验证码，使用封装的显式等待函数等待滑块验证消失
                    resolved, waited_seconds = waitCaptchaResolved(driver)
                    if not resolved:
                        print(f'  验证码等待超时，跳过该商家')
                        try:
                            driver.close()
                            driver.switch_to.window(main_window)
                        except Exception:
                            pass
                        continue
                    # 若确实经历过滑块验证，提示一次，并在验证通过后重新获取页面文本
                    if waited_seconds > 0:
                        print(f'  验证通过，继续采集...')
                        # 重新读取完成验证后的页面文本，避免读取到验证前的内容
                        try:
                            page_text = driver.find_element(By.TAG_NAME, 'body').text
                        except Exception:
                            page_text = page_text

                    # 先尝试从 DOM 提取联系人、电话、手机、传真、地址，再用正则兜底
                    member_name, tel, mobile, fax, address = extractContactByDom(driver)
                    if not member_name or not tel or not mobile or not fax or not address:
                        r_name, r_tel, r_mobile, r_fax, r_addr = extractContactByRegex(page_text)
                        if not member_name:
                            member_name = r_name
                        if not tel:
                            tel = r_tel
                        if not mobile:
                            mobile = r_mobile
                        if not fax:
                            fax = r_fax
                        if not address:
                            address = r_addr

                    # 清洗地址：去掉误抓的「技术支持:旺铺管理」等后缀
                    address = cleanAddress(address or '')

                    address_preview = (address or '')[:30]
                    # print(f'  联系人: {member_name or "(无)"} | 电话: {tel or "(无)"} | 手机: {mobile or "(无)"} | 传真: {fax or "(无)"} | 地址: {address_preview or "(无)"}')
                    # 仅在有手机号时才写入 Excel（表头：企业名称、当前城市、联系方式、联系人、手机、地址）
                    mobile_stripped = (mobile or '').strip()
                    if mobile_stripped:
                        # 当前城市列：优先写入 TARGET_CITY；若为空则写入省份或「全国」
                        current_city_for_excel = TARGET_CITY or (TARGET_REGION or '全国')
                        row_data = (
                            title_value,
                            current_city_for_excel,
                            contact_url,
                            member_name or '',
                            mobile_stripped,
                            address or '',
                        )
                        for col, val in enumerate(row_data, start=1):
                            ws.cell(row=excel_row, column=col, value=val)
                        excel_row += 1
                        wb.save(OUTPUT_EXCEL)

                        # 更新全局已采集数量，并根据 TOTAL_MAX_SHOPS 判断是否需要提前结束本次采集
                        total_shops_collected += 1
                        # 每次成功写入一条数据后，打印当前已写入的总数量，方便观察采集进度（立即刷新输出）
                        print(f'  本次 Excel 已累计写入 {total_shops_collected} 条数据', flush=True)
                        if TOTAL_MAX_SHOPS > 0 and total_shops_collected >= TOTAL_MAX_SHOPS:
                            print(f'本次已采集商家数量达到上限 {TOTAL_MAX_SHOPS}，停止后续采集')
                            stop_collecting = True

                        # 根据开始时间计算本条已耗时，不足目标时长则额外等待一段时间，避免访问过快触发风控
                        elapsed = time.time() - shop_start_ts
                        target_seconds = MIN_SECONDS_PER_SHOP + random.uniform(-2, 2)
                        if target_seconds < MIN_SECONDS_PER_SHOP * 0.8:
                            target_seconds = MIN_SECONDS_PER_SHOP * 0.8
                        extra_sleep = max(0, target_seconds - elapsed)
                        if extra_sleep > 0:
                            time.sleep(extra_sleep)
                    else:
                        print('  无手机号，跳过写入 Excel')
                    # 关闭联系方式页标签，切回主标签（搜索结果页），继续本页下一个或翻页
                    driver.close()
                    driver.switch_to.window(main_window)

                except Exception as e:
                    print(f'  采集失败: {e}')
                    try:
                        if driver.current_window_handle != main_window:
                            driver.close()
                        driver.switch_to.window(main_window)
                        time.sleep(3)
                    except Exception:
                        pass
                    continue
            # 主标签仍在搜索结果页：先判断是否有下一页，没有则直接结束当前城市的采集
            page_elements = driver.find_elements(By.CSS_SELECTOR, "a.fui-next")
            if not page_elements:
                print(f'城市 {TARGET_CITY or "整省/全国"} 没有下一页，结束该城市采集')
                break
            next_btn = page_elements[0]
            # 检查“下一页”按钮是否处于禁用状态（已是最后一页）
            try:
                next_class = (next_btn.get_attribute('class') or '').lower()
                aria_disabled = (next_btn.get_attribute('aria-disabled') or '').lower()
                if 'disabled' in next_class or 'fui-next-disabled' in next_class or aria_disabled == 'true':
                    print(f'城市 {TARGET_CITY or "整省/全国"} 已是最后一页，结束该城市采集')
                    break
            except Exception:
                pass
            # 有下一页时才滚动并点击下一页，避免无谓等待
            time.sleep(2)
            scrollToBottom(driver)
            time.sleep(2)
            try:
                closeKnownPopups(driver)
            except Exception:
                pass
            driver.execute_script("arguments[0].click();", next_btn)
            time.sleep(3)
            page_num += 1
        except Exception as e:
            print('error:', e)
            page_error_count += 1
            # 若连续页面级异常次数超过上限，则主动中止当前城市的循环，避免极端情况下无限重试
            if page_error_count >= MAX_PAGE_ERRORS:
                print(f'城市 {TARGET_CITY or "整省/全国"} 页面级异常次数已达到上限 {MAX_PAGE_ERRORS} 次，停止该城市采集循环')
                break
            continue

    # 若已达到全局采集上限，则不再进入后续城市循环
    if TOTAL_MAX_SHOPS > 0 and total_shops_collected >= TOTAL_MAX_SHOPS:
        break

# 所有城市采集完成或达到上限后：保存数据并关闭浏览器
try:
    wb.save(OUTPUT_EXCEL)
except Exception as e:
    print(f'保存 Excel 时出错: {e}')
print(f'采集完成！数据已保存到 {OUTPUT_EXCEL}', flush=True)
# 采集结束后，打印本次写入 Excel 的数据总条数，便于确认结果数量（立即刷新输出）
print(f'本次共向 Excel 写入 {total_shops_collected} 条数据', flush=True)
try:
    driver.quit()
except Exception:
    pass
