#! /usr/bin/env python
# coding:utf-8

from __future__ import annotations

import argparse
import io
import random
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from multiprocessing import freeze_support
from typing import Optional
from urllib.parse import quote, urlparse

from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import undetected_chromedriver as uc

# 统一 stdin/stdout 为 UTF-8，避免 Windows 下粘贴中文乱码或无法输入
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stdin, 'buffer'):
    sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8', errors='replace')

# ── 联系信息提取相关常量 ─────────────────────────────────
LABEL_CONTACT = '联系人'
LABEL_TEL = '电话'
LABEL_MOBILE = '手机'
LABEL_FAX = '传真'
LABEL_ADDRESS = '地址'

PATTERN_CONTACT_LABEL = re.compile(r'联系人[：:]\s*([^\n]{1,20})')
PATTERN_CONTACT_SUFFIX = re.compile(r'([\u4e00-\u9fa5]{2,5})(?:先生|女士|经理|总监|总裁|老板)')
PATTERN_TEL = re.compile(r'电话[：:]\s*([0-9\-\u00a0\s]+)')
PATTERN_MOBILE = re.compile(r'手机[：:]\s*([0-9\-\u00a0\s]+)')
PATTERN_FAX = re.compile(r'传真[：:]\s*([0-9\-\u00a0\s]+)')
PATTERN_ADDRESS = re.compile(
    r'地址[：:]\s*([^\n]+?)(?=\s*$|邮编|传真|公司名称|邮箱|联系人|电话|手机|技术支持)',
    re.DOTALL,
)
PATTERN_ADDRESS_SIMPLE = re.compile(r'地址[：:]\s*([^\n]+)')
PATTERN_ADDRESS_NOISE = re.compile(r'\s*技术支持[：:][^\n]*$')

MAX_ADDRESS_LEN = 200

# ── 搜索与分页常量 ──────────────────────────────────────
RESOLVE_REDIRECT_HOST = 'dj.1688.com'
RESOLVE_REDIRECT_PATH = 'ci_bb'
SPM_ANCHOR_PREFIX = 'a2615.'
OUTPUT_EXCEL_PREFIX = 'data'

DEFAULT_SEARCH_KEYWORDS: list[str] = [
    '机械设备', '精密加工', '五金工具', '工业耗材', '汽车配件', 'led照明', '家具',
    '家装建材', '塑料包装', '印刷', '健身器材', '户外园艺', '电子元器件', '传感器', '安防设备', '宠物用品', '原材料',
]
DEFAULT_TARGET_REGION = '广东'
DEFAULT_PROVINCE_CITY_MAP: dict[str, list[str]] = {
    '广东': ['东莞'],
}

MAX_FETCH_PER_PAGE = 0
EXCEL_HEADERS = ('企业名称', '关键词', '当前城市', '联系方式', '联系人', '手机', '地址')

VERBOSE_LOG = False


def logVerbose(message: str) -> None:
    """详细日志输出：仅在 VERBOSE_LOG 为 True 时打印。"""
    if VERBOSE_LOG:
        print(message)

# ── 运行参数默认值 ───────────────────────────────────────
DEFAULT_CAPTCHA_WAIT_TIMEOUT = 60
DEFAULT_MAX_PAGE_ERRORS = 10
DEFAULT_TOTAL_MAX_SHOPS = 0

# ── 节奏控制（反爬：随机化、冷却）────────────────────────
MIN_SECONDS_PER_SHOP = 15
MIN_SECONDS_PER_SHOP_JITTER = 2  # 单条采集耗时在 MIN ± JITTER 内随机
PAGE_TURN_WAIT_MIN = 3
PAGE_TURN_WAIT_MAX = 7
# 任务切换：每个新任务开始前随机等待，避免连续请求
TASK_START_DELAY_MIN = 2
TASK_START_DELAY_MAX = 6
# 打开新页面/新标签后等待，模拟用户阅读
PAGE_OPEN_DELAY_MIN = 1
PAGE_OPEN_DELAY_MAX = 3
# 无手机号时仍保持一定间隔，避免空结果也高频请求
SHOP_NO_MOBILE_WAIT_MIN = 3
SHOP_NO_MOBILE_WAIT_MAX = 6
# 每采集 N 个商家后休息一段时间，降低连续请求特征
REST_EVERY_N_SHOPS = 8
REST_DURATION_MIN = 25
REST_DURATION_MAX = 45
# 访问被拒/验证码通过后的“恢复期”额外等待
RECOVERY_AFTER_DENIED_EXTRA_MIN = 5
RECOVERY_AFTER_DENIED_EXTRA_MAX = 12
# 登录后建立 1688 会话的等待
LOGIN_TO_1688_WAIT_MIN = 3
LOGIN_TO_1688_WAIT_MAX = 6

# ── 弹窗处理 JS ────────────────────────────────────────
JS_REMOVE_BAXIA_MASK = (
    "var d=document.querySelector('.baxia-dialog');"
    "var c=d&&(d.innerText||'').match(/拖动|验证|slide/i);"
    "var m=document.querySelector('.baxia-dialog-mask');"
    "if(m&&!c)m.remove();"
)
JS_REMOVE_BAXIA_DIALOG = (
    "var d=document.querySelector('.baxia-dialog');"
    "var c=d&&(d.innerText||'').match(/拖动|验证|slide/i);"
    "if(d&&!c)d.remove();"
)
TEXT_ACCESS_DENIED = '访问被拒绝'
MAX_ACCESS_DENIED_CLOSE_ATTEMPTS = 3
ACCESS_DENIED_COOLDOWN_MIN = 15
ACCESS_DENIED_COOLDOWN_MAX = 30

# ── 滑块自动拖拽相关常量 ──────────────────────────────────
SLIDER_BTN_SELECTORS = [
    '#nc_1_n1z',
    '.nc_iconfont.btn_slide',
    '.btn_slide',
    '#nc_1__scale_text .btn_slide',
    '.nc-lang-cnt .btn_slide',
    '.slide-btn',
    '#slide-btn',
    'span.nc_iconfont',
    '.baxia-dialog .btn_slide',
    '.baxia-dialog .nc_iconfont',
]
SLIDER_TRACK_SELECTORS = [
    '#nc_1__scale_text',
    '#nc_1_wrapper',
    '.nc-lang-cnt',
    '.scale_text',
    '.slide-track',
    '#slide-track',
    '.baxia-dialog .nc-lang-cnt',
    '.baxia-dialog .scale_text',
]
SLIDER_IFRAME_SELECTORS = [
    '#baxia-dialog-content iframe',
    '.baxia-dialog iframe',
    '#sufei-dialog-content iframe',
    'iframe[src*="nocaptcha"]',
    'iframe[src*="captcha"]',
]
SLIDER_DEFAULT_DISTANCE = 340
SLIDER_MAX_RETRY = 3
SLIDER_STEP_MIN_MS = 8
SLIDER_STEP_MAX_MS = 25

# ── 滚动加载常量（反爬：随机化步长与间隔）────────────────
SCROLL_STEP_PX = 600
SCROLL_STEP_PX_JITTER = 80   # 每步滚动量 ±jitter 随机
SCROLL_STEP_WAIT_MIN = 0.4
SCROLL_STEP_WAIT_MAX = 0.9
SCROLL_FINAL_WAIT = 1.5
EXPECTED_ITEMS_PER_PAGE = 20

# ── 等待超时 ────────────────────────────────────────────
LOGIN_WAIT_TIMEOUT = 120
CONTACT_PAGE_WAIT_TIMEOUT = 15


# ═══════════════════════════════════════════════════════════
#  数据类：运行时配置
# ═══════════════════════════════════════════════════════════

@dataclass
class ScraperConfig:
    """运行时配置，聚合命令行参数和用户交互输入。"""
    captcha_timeout: int = DEFAULT_CAPTCHA_WAIT_TIMEOUT
    max_page_errors: int = DEFAULT_MAX_PAGE_ERRORS
    total_max_shops: int = DEFAULT_TOTAL_MAX_SHOPS
    keywords: list[str] = field(default_factory=lambda: list(DEFAULT_SEARCH_KEYWORDS))
    target_region: str = DEFAULT_TARGET_REGION
    province_city_map: dict[str, list[str]] = field(default_factory=lambda: dict(DEFAULT_PROVINCE_CITY_MAP))


# ═══════════════════════════════════════════════════════════
#  纯函数 / 工具函数（不依赖 driver）
# ═══════════════════════════════════════════════════════════

def cleanAddress(raw: str) -> str:
    """清洗地址字符串：去掉末尾「技术支持:xxx」等噪音，截断到最大长度。"""
    if not (raw or '').strip():
        return ''
    cleaned = re.sub(PATTERN_ADDRESS_NOISE, '', raw).strip()
    return cleaned[:MAX_ADDRESS_LEN] if cleaned else ''


def getShopOrigin(shop_url: str) -> str:
    """从任意店铺 URL 解析出「协议 + 域名」，用于拼接 /page/contactinfo.htm。"""
    if not (shop_url or '').strip():
        return ''
    try:
        parsed = urlparse(shop_url.strip())
        if parsed.scheme and parsed.netloc:
            return f'{parsed.scheme}://{parsed.netloc}'
    except Exception:
        pass
    return shop_url.rstrip('/').split('/page/')[0].split('?')[0] or shop_url


def buildSearchUrl(keyword: str, province_name: str, city_name: str) -> str:
    """根据关键词、省份与城市构造 1688 公司搜索页 URL（GBK 编码）。"""
    try:
        base = (
            'https://s.1688.com/company/company_search.htm?'
            'keywords=' + quote(keyword, encoding='gbk', safe='')
            + '&n=y&spm=a260k.635.1998096057.d1'
        )
        if province_name:
            base += '&province=' + quote(province_name, encoding='gbk', safe='')
            if city_name:
                base += '&city=' + quote(city_name, encoding='gbk', safe='')
        return base
    except Exception:
        return (
            'https://s.1688.com/company/company_search.htm?'
            'keywords=' + quote(keyword, encoding='gbk', safe='')
            + '&n=y&spm=a260k.635.1998096057.d1'
        )


def buildOutputFileName(config: ScraperConfig) -> str:
    """
    根据当前日期时间、搜索关键词和地区构造导出 Excel 文件名。
    格式：YYYYMMDD_HHMMSS_搜索内容_地区.xlsx
    """
    MAX_SHOW_KEYWORDS = 3
    try:
        now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        region_desc = config.target_region or '全国'

        if config.keywords:
            shown = '+'.join(config.keywords[:MAX_SHOW_KEYWORDS])
            if len(config.keywords) > MAX_SHOW_KEYWORDS:
                keywords_part = f'{shown}等{len(config.keywords)}个'
            else:
                keywords_part = shown
        else:
            keywords_part = '未命名'

        return f'{now_str}_{keywords_part.replace(" ", "")}_{region_desc}.xlsx'
    except Exception:
        return f'{OUTPUT_EXCEL_PREFIX}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'


def extractContactByRegex(page_text: str) -> tuple[str, str, str, str, str]:
    """
    从页面纯文本中用正则提取联系人、电话、手机、传真、地址。
    用于 DOM 取不到时的兜底。返回 (联系人, 电话, 手机, 传真, 地址)。
    """
    if not (page_text or '').strip():
        return ('', '', '', '', '')

    member_name = ''
    tel = ''
    mobile = ''
    fax = ''
    address = ''

    m = PATTERN_CONTACT_LABEL.search(page_text)
    if m:
        member_name = re.sub(r'[\s\d].*', '', m.group(1).strip())[:20]
    if not member_name:
        m = PATTERN_CONTACT_SUFFIX.search(page_text)
        if m:
            member_name = m.group(1).strip()[:30]

    m = PATTERN_TEL.search(page_text)
    if m:
        tel = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]

    m = PATTERN_MOBILE.search(page_text)
    if m:
        mobile = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]

    m = PATTERN_FAX.search(page_text)
    if m:
        fax = re.sub(r'[\s\u00a0]+', ' ', re.sub(r'[^\d\-]', '', m.group(1))).strip()[:50]

    m = PATTERN_ADDRESS.search(page_text)
    if m:
        address = m.group(1).strip()
    if not address:
        m = PATTERN_ADDRESS_SIMPLE.search(page_text)
        if m:
            address = m.group(1).strip()
    address = address[:MAX_ADDRESS_LEN]

    return (member_name, tel, mobile, fax, address)


def randomDelay(min_sec: float, max_sec: float) -> None:
    """
    反爬：在 [min_sec, max_sec] 内随机休眠，统一人性化延迟。
    避免固定间隔被识别为脚本。
    """
    if max_sec <= 0:
        return
    delay = random.uniform(max(0, min_sec), max_sec)
    if delay > 0:
        time.sleep(delay)


# ═══════════════════════════════════════════════════════════
#  配置加载与用户交互（在创建 Scraper 前调用）
# ═══════════════════════════════════════════════════════════

def loadRuntimeConfig() -> ScraperConfig:
    """通过命令行参数加载运行配置，返回 ScraperConfig 实例。"""
    parser = argparse.ArgumentParser(description='1688 商家联系方式采集脚本运行参数')
    parser.add_argument(
        '--captcha-timeout', type=int, default=DEFAULT_CAPTCHA_WAIT_TIMEOUT,
        help='滑块验证码最长等待时间（秒），默认 60 秒',
    )
    parser.add_argument(
        '--max-page-errors', type=int, default=DEFAULT_MAX_PAGE_ERRORS,
        help='分页循环中允许的最大页面级异常次数，默认 10 次',
    )
    parser.add_argument(
        '--max-shops', type=int, default=DEFAULT_TOTAL_MAX_SHOPS,
        help='本次运行最多采集的商家数量，0 或负数表示不限制',
    )
    args, _ = parser.parse_known_args()

    return ScraperConfig(
        captcha_timeout=args.captcha_timeout if args.captcha_timeout > 0 else DEFAULT_CAPTCHA_WAIT_TIMEOUT,
        max_page_errors=args.max_page_errors if args.max_page_errors > 0 else DEFAULT_MAX_PAGE_ERRORS,
        total_max_shops=args.max_shops if args.max_shops >= 0 else DEFAULT_TOTAL_MAX_SHOPS,
    )


def collectUserInput(config: ScraperConfig) -> ScraperConfig:
    """交互式收集用户输入：搜索关键词、目标省份、指定城市。直接回车使用默认值。"""
    # ── 输入搜索关键词 ──
    default_kw_str = '、'.join(DEFAULT_SEARCH_KEYWORDS)
    print(f'当前默认关键词（共 {len(DEFAULT_SEARCH_KEYWORDS)} 个）:')
    print(f'  {default_kw_str}')
    print()
    kw_input = input('请输入搜索关键词（多个用逗号分隔，直接回车使用默认值）: ').strip()
    if kw_input:
        raw_list = re.split(r'[,，、]+', kw_input)
        config.keywords = [kw.strip() for kw in raw_list if kw.strip()]
    else:
        config.keywords = list(DEFAULT_SEARCH_KEYWORDS)
    print(f'  → 本次关键词（{len(config.keywords)} 个）: {config.keywords}')
    print()

    # ── 输入目标省份 ──
    print(f'当前默认省份: {DEFAULT_TARGET_REGION or "（不限省份）"}')
    region_input = input('请输入目标省份（如 广东、浙江，留空不限，直接回车使用默认值）: ').strip()
    config.target_region = region_input if region_input else DEFAULT_TARGET_REGION
    print(f'  → 本次省份: {config.target_region or "（不限省份，全国范围）"}')
    print()

    # ── 输入指定城市 ──
    default_cities = DEFAULT_PROVINCE_CITY_MAP.get(config.target_region, [])
    if default_cities:
        print(f'当前默认城市: {", ".join(default_cities)}')
    else:
        print('当前无预置城市（将自动从页面解析）')
    city_input = input(
        '请输入指定城市（多个用逗号分隔，留空自动解析，直接回车使用默认值）: '
    ).strip()
    if city_input:
        raw_cities = re.split(r'[,，、]+', city_input)
        user_cities = [c.strip() for c in raw_cities if c.strip()]
        if user_cities and config.target_region:
            config.province_city_map = {config.target_region: user_cities}
        else:
            config.province_city_map = {}
    elif default_cities:
        config.province_city_map = dict(DEFAULT_PROVINCE_CITY_MAP)
    else:
        config.province_city_map = {}

    final_cities = config.province_city_map.get(config.target_region, [])
    if final_cities:
        print(f'  → 本次城市: {", ".join(final_cities)}')
    else:
        print('  → 本次城市: （自动从页面解析全省城市列表）')

    print()
    print('=' * 60)
    print(f'  关键词: {config.keywords}')
    print(f'  省份:   {config.target_region or "不限"}')
    print(f'  城市:   {final_cities or "自动解析"}')
    print('=' * 60)
    print()
    confirm = input('确认以上配置并开始采集？（回车确认 / 输入 n 退出）: ').strip().lower()
    if confirm in ('n', 'no', '否'):
        print('已取消，程序退出。')
        sys.exit(0)
    print()
    return config


# ═══════════════════════════════════════════════════════════
#  核心采集器类
# ═══════════════════════════════════════════════════════════

class AlibabaScraper:
    """1688 商家联系方式采集器，封装浏览器驱动、Excel 写入和采集流程。"""

    def __init__(self, config: ScraperConfig):
        self.config = config
        self.driver: Optional[webdriver.Chrome] = None
        self.wait: Optional[WebDriverWait] = None
        self.wb: Optional[Workbook] = None
        self.ws = None
        self.output_file = ''
        self.excel_row = 2
        self.seen_shops: set[str] = set()
        self.total_collected = 0
        self._stop = False
        # 反爬：访问被拒/验证码后的恢复期，下一批操作延长等待
        self._recovery_until: float = 0

    # ── 生命周期 ──────────────────────────────────────────

    def __enter__(self):
        self._initDriver()
        self._initExcel()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    def close(self):
        """保存 Excel 并关闭浏览器，确保资源释放。"""
        try:
            if self.wb:
                self.wb.save(self.output_file)
                print(f'数据已保存到 {self.output_file}')
        except Exception as e:
            print(f'✗ 保存 Excel 时出错: {e}')
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass

    # ── 初始化 ────────────────────────────────────────────

    def _initDriver(self):
        """
        优先使用 undetected-chromedriver 创建浏览器实例（底层 patch 反检测），
        不可用时依次退化到普通 Selenium Chrome、Firefox。
        反爬：禁用自动化特征、设置常见窗口尺寸、注入反检测脚本。
        """
        # 常见窗口尺寸（反爬：避免无头/默认尺寸被识别）
        width = random.randint(1280, 1440)
        height = random.randint(720, 900)

        # 方案 1：undetected-chromedriver（推荐）
        try:
            options = uc.ChromeOptions()
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--lang=zh-CN')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--no-sandbox')
            self.driver = uc.Chrome(options=options)
            self.driver.set_window_size(width, height)
            self.wait = WebDriverWait(self.driver, LOGIN_WAIT_TIMEOUT)
            return
        except Exception as e_uc:
            print(f'undetected-chromedriver 启动失败: {str(e_uc)[:80]}')

        # 方案 2：普通 Selenium Chrome + 手动反检测
        try:
            options = webdriver.ChromeOptions()
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
            options.add_experimental_option('useAutomationExtension', False)
            self.driver = webdriver.Chrome(options=options)
            self.driver.set_window_size(width, height)
            # 反爬：注入反检测脚本，隐藏 webdriver、统一语言与插件特征
            self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': """
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'languages', { get: () => ['zh-CN', 'zh', 'en'] });
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                if (!window.chrome) {
                    window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };
                }
                const originalQuery = window.navigator.permissions.query;
                window.navigator.permissions.query = (parameters) => (
                    parameters.name === 'notifications'
                        ? Promise.resolve({ state: Notification.permission })
                        : originalQuery(parameters)
                );
            """})
            self.wait = WebDriverWait(self.driver, LOGIN_WAIT_TIMEOUT)
            return
        except Exception as e_chrome:
            print(f'普通 Chrome 不可用: {str(e_chrome)[:50]}')

        # 方案 3：Firefox 兜底
        try:
            self.driver = webdriver.Firefox()
            self.driver.set_window_size(width, height)
            self.wait = WebDriverWait(self.driver, LOGIN_WAIT_TIMEOUT)
        except Exception as e_ff:
            raise RuntimeError(f'所有浏览器均不可用: {e_ff}') from e_ff

    def _initExcel(self):
        """创建 Excel 工作簿并写入表头。"""
        self.output_file = buildOutputFileName(self.config)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '采集数据'
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            self.ws.cell(row=1, column=col, value=header)

    # ── 登录流程 ──────────────────────────────────────────

    def login(self):
        """打开淘宝登录页并等待用户扫码，登录后访问 1688 建立会话。"""
        self.driver.get('https://login.taobao.com/member/login.jhtml')

        try:
            self.wait.until(EC.url_contains('taobao.com/'))
            self.wait.until_not(EC.url_contains('login.taobao.com'))
            print('✓ 登录成功！')
        except Exception:
            if 'login.taobao.com' not in self.driver.current_url:
                print('✓ 登录成功！')
            else:
                print('✗ 登录超时，请重新运行脚本并及时扫码')
                raise SystemExit(1)
        self.driver.get('https://www.1688.com/')
        randomDelay(LOGIN_TO_1688_WAIT_MIN, LOGIN_TO_1688_WAIT_MAX)

    # ── 主采集入口 ────────────────────────────────────────

    def run(self):
        """主采集流程：解析城市列表 → 生成任务队列 → 逐任务采集。"""
        city_list = self._resolveCityList()
        task_queue = [
            (city, kw)
            for city in city_list
            for kw in self.config.keywords
        ]
        print(f'共 {len(task_queue)} 个采集任务（{len(city_list)} 城市 × {len(self.config.keywords)} 关键词）')

        for idx, (city, keyword) in enumerate(task_queue, start=1):
            if self._stop:
                break
            self._processTask(idx, len(task_queue), city, keyword)

        print(f'✓ 采集完成！共写入 {self.total_collected} 条数据 → {self.output_file}')

    # ── 城市列表解析 ──────────────────────────────────────

    def _resolveCityList(self) -> list[str]:
        """根据省份配置解析城市列表，无省份则返回 ['']（全国维度抓取一次）。"""
        if not self.config.target_region:
            return ['']

        city_list = self._getCityListByProvince(
            self.config.target_region, self.config.keywords[0]
        )
        if not city_list:
            print(f'未能解析省份「{self.config.target_region}」的城市列表，将按整省抓取')
            return ['']

        effective = [c for c in city_list if c]
        print(f'省份「{self.config.target_region}」共 {len(effective)} 个城市: {city_list}')
        return city_list

    def _getCityListByProvince(self, province_name: str, keyword: str) -> list[str]:
        """
        从 1688 搜索页「所在地区」筛选中解析指定省份下的城市列表。
        按视觉顺序（从上到下、从左到右）排序返回。
        """
        # 优先使用预置城市列表
        preset = self.config.province_city_map.get(province_name)
        if preset:
            logVerbose(f'省份「{province_name}」使用预置城市列表: {preset}')
            return list(preset)

        city_list: list[str] = []
        try:
            url = buildSearchUrl(keyword, province_name, '')
            self.driver.get(url)

            try:
                WebDriverWait(self.driver, 20).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, 'div.sm-widget-address, div.sm-widget-region, div.address-widget')
                    )
                )
            except Exception:
                pass

            # 尝试点击省份展开城市列表
            self._clickProvinceExpander(province_name)

            candidates = self._collectCityCandidates(province_name)
            if not candidates:
                print(f'未在省份「{province_name}」页面上找到城市链接元素')
                return []

            # 按视觉位置排序（从上到下、从左到右），然后去重
            candidates.sort(key=lambda item: (round(item[0], 1), round(item[1], 1)))
            seen: set[str] = set()
            for _, _, name in candidates:
                if name not in seen:
                    seen.add(name)
                    city_list.append(name)
        except Exception as e:
            print(f'自动解析省份「{province_name}」城市列表失败: {e}')

        return city_list

    def _clickProvinceExpander(self, province_name: str):
        """尝试点击省份元素以展开城市列表。"""
        try:
            elements = self.driver.find_elements(By.XPATH, f"//*[text()='{province_name}']")
            for el in elements:
                try:
                    self.driver.execute_script('arguments[0].click();', el)
                    time.sleep(1)
                    break
                except Exception:
                    continue
        except Exception:
            pass

    def _collectCityCandidates(self, province_name: str) -> list[tuple[float, float, str]]:
        """
        从地区筛选区域收集所有城市候选元素及其视觉位置。
        返回 [(top, left, city_name), ...] 列表。
        """
        candidates: list[tuple[float, float, str]] = []
        IGNORED_NAMES = {'不限', '全部', '全国'}
        CITY_PATTERN = re.compile(r'^[\u4e00-\u9fa5]{1,6}$')

        def _extractFromLinks(links):
            for link in links:
                text_val = (link.text or '').strip()
                if not text_val or text_val in IGNORED_NAMES or not CITY_PATTERN.match(text_val):
                    continue
                top, left = self._getElementPosition(link)
                candidates.append((top, left, text_val))

        # 方式 1：从地区筛选容器内查找
        try:
            containers = self.driver.find_elements(
                By.XPATH,
                "//div[contains(@class,'sm-widget-address') or "
                "contains(@class,'sm-widget-region') or "
                "contains(@class,'address-widget')]"
            )
            for container in containers:
                _extractFromLinks(container.find_elements(By.TAG_NAME, 'a'))
        except Exception:
            pass

        # 方式 2：全局兜底查找带 &city= 参数的链接
        if not candidates:
            try:
                links = self.driver.find_elements(
                    By.XPATH,
                    "//a[@href and contains(@href,'company_search.htm') and contains(@href,'city=')]"
                )
                _extractFromLinks(links)
            except Exception:
                pass

        return candidates

    def _getElementPosition(self, element) -> tuple[float, float]:
        """获取 DOM 元素在视口中的位置 (top, left)。"""
        try:
            rect = self.driver.execute_script(
                'var r = arguments[0].getBoundingClientRect(); return [r.top, r.left];',
                element,
            )
            return (float(rect[0]), float(rect[1])) if rect and len(rect) >= 2 else (0.0, 0.0)
        except Exception:
            return (0.0, 0.0)

    # ── 单个任务（城市 + 关键词）处理 ────────────────────────

    def _processTask(self, task_idx: int, total_tasks: int, city: str, keyword: str):
        """处理单个（城市, 关键词）任务：逐页采集直到无下一页或达到上限。"""
        city_desc = city or '整省/全国'
        logVerbose('━' * 50)
        logVerbose(f'任务 [{task_idx}/{total_tasks}] 关键词「{keyword}」城市「{city_desc}」')

        # 反爬：新任务开始前随机等待，避免连续请求
        randomDelay(TASK_START_DELAY_MIN, TASK_START_DELAY_MAX)
        search_url = buildSearchUrl(keyword, self.config.target_region, city)
        self.driver.get(search_url)
        randomDelay(PAGE_OPEN_DELAY_MIN, PAGE_OPEN_DELAY_MAX)

        main_window = self.driver.current_window_handle
        page_error_count = 0
        page_num = 1

        while not self._stop:
            try:
                should_continue = self._processPage(
                    keyword, city, page_num, main_window
                )
                if not should_continue:
                    break

                # 翻页
                if not self._goToNextPage(keyword, city, page_num):
                    break
                page_num += 1

            except Exception as e:
                print(f'页面级异常: {e}')
                page_error_count += 1
                if page_error_count >= self.config.max_page_errors:
                    print(f'任务「{keyword}」「{city_desc}」页面异常达上限 {self.config.max_page_errors} 次，终止该任务')
                    break

    def _processPage(
        self, keyword: str, city: str, page_num: int, main_window: str
    ) -> bool:
        """
        处理当前搜索结果页：抓取所有商家的联系信息。
        返回 True 表示可继续翻页，False 表示应终止当前任务。
        """
        city_desc = city or '整省/全国'
        self._maybeRecoveryWait()
        self._closeKnownPopups()
        if self._closeAccessDeniedPopup():
            print('检测到「访问被拒绝」弹窗，已尝试关闭')

        # 等待搜索结果列表加载
        try:
            self.wait.until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.company-name'))
            )
        except Exception:
            pass

        self._scrollToLoadAllResults()

        titles = self.driver.find_elements(By.CSS_SELECTOR, 'a.company-name')
        if not titles:
            self._logEmptyPage(page_num)
            return False

        logVerbose(f'关键词「{keyword}」城市「{city_desc}」第 {page_num} 页找到 {len(titles)} 个商家')

        # 检查是否已到末页
        if self._isLastPage(keyword, city_desc, page_num):
            return False

        # 收集当前页所有商家数据并按垂直位置排序
        page_data = self._collectPageData(titles)
        ordered = page_data[:MAX_FETCH_PER_PAGE] if MAX_FETCH_PER_PAGE else page_data

        for _, title_el, title_value, href_value in ordered:
            if self._stop:
                break
            self._scrollToElement(title_el)
            self._processShop(title_el, title_value, href_value, keyword, city, main_window)

        return True

    def _collectPageData(self, title_elements) -> list[tuple[float, object, str, str]]:
        """收集当前页所有商家信息并按垂直位置排序，返回 [(top, element, title, href), ...]。"""
        page_data: list[tuple[float, object, str, str]] = []
        for el in title_elements:
            title_value = el.get_attribute('title') or el.text
            href_value = el.get_attribute('href')
            if not href_value:
                continue
            top = self._getElementPosition(el)[0]
            page_data.append((top, el, title_value, href_value))

        page_data.sort(key=lambda x: x[0])
        return page_data

    def _logEmptyPage(self, page_num: int):
        """列表为空时记录日志，区分「访问被拒绝」和「无结果」两种情况。"""
        try:
            body_text = self.driver.find_element(By.TAG_NAME, 'body').text or ''
            if TEXT_ACCESS_DENIED in body_text:
                print(f'第 {page_num} 页列表为空，存在「访问被拒绝」提示')
            else:
                print('该关键词在当前城市没有搜索结果，跳到下一个任务')
        except Exception:
            pass

    def _isLastPage(self, keyword: str, city_desc: str, page_num: int) -> bool:
        """解析分页控件，判断当前是否已达末页。"""
        try:
            pager_elements = self.driver.find_elements(
                By.CSS_SELECTOR,
                "div.fui-pager, div.sm-pagination, div[class*='pagination'], span.page-count"
            )
            pager_text = ''
            for el in pager_elements:
                text_val = (el.text or '').strip()
                if text_val:
                    pager_text = text_val
                    break

            total_pages = 0
            if pager_text:
                m = re.search(r'/\s*(\d+)', pager_text) or re.search(r'共\s*(\d+)\s*页', pager_text)
                if m:
                    total_pages = int(m.group(1))

            if total_pages > 0:
                logVerbose(f'关键词「{keyword}」城市「{city_desc}」搜索结果共 {total_pages} 页')
                if page_num >= total_pages:
                    logVerbose(f'已达末页（第 {page_num}/{total_pages} 页），结束该任务')
                    return True
        except Exception:
            pass
        return False

    # ── 单个商家处理 ──────────────────────────────────────

    def _processShop(
        self, title_el, title_value: str, href_value: str,
        keyword: str, city: str, main_window: str,
    ):
        """在新标签中打开商家联系页，提取联系信息，写入 Excel。"""
        self._maybeRecoveryWait()
        shop_start = time.time()
        href = (href_value or '').strip()
        is_redirect = (RESOLVE_REDIRECT_HOST in href) or (RESOLVE_REDIRECT_PATH in href)

        try:
            shop_origin, contact_url = self._openContactPage(href, is_redirect, main_window)
            if shop_origin is None:
                return

            dedup_key = shop_origin or href or title_value
            if dedup_key in self.seen_shops:
                logVerbose(f'  跳过重复店铺（已采集）: {title_value}')
                self._closeTabAndReturn(main_window)
                return
            self.seen_shops.add(dedup_key)

            # 确保真正进入联系页
            self._ensureContactPage(contact_url)

            # 等待联系页内容渲染
            page_text = self._waitForContactContent()

            # 处理验证码
            resolved, waited = self._waitCaptchaResolved()
            if not resolved:
                print(f'验证码超时，跳过: {title_value}')
                self._closeTabAndReturn(main_window)
                return

            if waited > 0:
                print('验证通过，刷新页面重新获取联系数据...')
                # 反爬：验证码通过后短暂恢复期，降低紧接着的请求频率
                self._recovery_until = time.time() + random.uniform(3, 8)
                page_text = self._refreshAndGetText(page_text)

            # 提取联系信息（DOM 优先，正则兜底）
            contact = self._extractContact(page_text)
            mobile_stripped = (contact['mobile'] or '').strip()

            if mobile_stripped:
                self._writeToExcel(title_value, keyword, city, contact_url, contact)
                self.total_collected += 1
                print(f'已累计写入 {self.total_collected} 条数据', flush=True)
                self._checkCollectLimit()
                # 反爬：每采集 N 个商家后休息一段时间
                if REST_EVERY_N_SHOPS > 0 and self.total_collected % REST_EVERY_N_SHOPS == 0:
                    rest_sec = random.uniform(REST_DURATION_MIN, REST_DURATION_MAX)
                    logVerbose(f'已采集 {self.total_collected} 条，休息 {rest_sec:.0f} 秒...')
                    time.sleep(rest_sec)
                self._throttle(shop_start)
            else:
                randomDelay(SHOP_NO_MOBILE_WAIT_MIN, SHOP_NO_MOBILE_WAIT_MAX)

            self.driver.close()
            self.driver.switch_to.window(main_window)

        except Exception as e:
            print(f'采集失败 [{title_value}]: {e}')
            self._safeCloseAndReturn(main_window)

    def _openContactPage(
        self, href: str, is_redirect: bool, main_window: str,
    ) -> tuple[Optional[str], str]:
        """
        在新标签中打开联系方式页。
        返回 (shop_origin, contact_url)；若应跳过该商家则返回 (None, '')。
        """
        if not is_redirect:
            shop_origin = getShopOrigin(href)
            if not shop_origin:
                print('  无法解析店铺域名，跳过')
                return None, ''
            contact_url = shop_origin.rstrip('/') + '/page/contactinfo.htm'
            self.driver.execute_script("window.open(arguments[0], '_blank');", contact_url)
            self.driver.switch_to.window(self.driver.window_handles[-1])
            randomDelay(PAGE_OPEN_DELAY_MIN, PAGE_OPEN_DELAY_MAX)
            self._closeKnownPopups()
        else:
            self.driver.execute_script("window.open(arguments[0], '_blank');", href)
            self.driver.switch_to.window(self.driver.window_handles[-1])
            randomDelay(PAGE_OPEN_DELAY_MIN, PAGE_OPEN_DELAY_MAX)
            self._closeKnownPopups()
            shop_origin = getShopOrigin((self.driver.current_url or '').strip())
            if not shop_origin:
                print('  无法解析店铺域名（跳转链接），跳过')
                self._closeTabAndReturn(main_window)
                return None, ''
            contact_url = shop_origin.rstrip('/') + '/page/contactinfo.htm'
            self.driver.get(contact_url)
            self._closeKnownPopups()

        return shop_origin, contact_url

    def _ensureContactPage(self, contact_url: str):
        """若未真正进入联系页，用当前域名再试一次。"""
        current = (self.driver.current_url or '').strip()
        if 'contactinfo' not in current and '.1688.com' in current:
            retry_origin = getShopOrigin(current)
            if retry_origin:
                self.driver.get(retry_origin.rstrip('/') + '/page/contactinfo.htm')
            self._closeKnownPopups()

        if 'contactinfo' not in (self.driver.current_url or ''):
            print(f'未进入联系方式页: {(self.driver.current_url or "")[:80]}')

    def _waitForContactContent(self) -> str:
        """等待联系方式页异步内容加载完成，返回页面文本。"""
        try:
            contact_wait = WebDriverWait(self.driver, CONTACT_PAGE_WAIT_TIMEOUT, poll_frequency=1)
            contact_wait.until(lambda d: any(
                kw in (d.find_element(By.TAG_NAME, 'body').text or '')
                for kw in ['电话', '手机', '地址', '联系人', '传真']
            ))
        except Exception:
            time.sleep(3)

        return self.driver.find_element(By.TAG_NAME, 'body').text or ''

    def _refreshAndGetText(self, fallback: str) -> str:
        """刷新页面并重新获取文本内容。"""
        try:
            self.driver.refresh()
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
            )
            return self.driver.find_element(By.TAG_NAME, 'body').text or ''
        except Exception:
            return fallback

    def _extractContact(self, page_text: str) -> dict[str, str]:
        """先 DOM 提取联系信息，不足部分用正则兜底。返回 dict。"""
        member_name, tel, mobile, fax, address = self._extractContactByDom()
        if not all([member_name, tel, mobile, fax, address]):
            r_name, r_tel, r_mobile, r_fax, r_addr = extractContactByRegex(page_text)
            member_name = member_name or r_name
            tel = tel or r_tel
            mobile = mobile or r_mobile
            fax = fax or r_fax
            address = address or r_addr

        return {
            'member_name': member_name,
            'tel': tel,
            'mobile': mobile,
            'fax': fax,
            'address': cleanAddress(address),
        }

    def _writeToExcel(
        self, title: str, keyword: str, city: str,
        contact_url: str, contact: dict[str, str],
    ):
        """将一条商家记录写入 Excel 并即时保存。"""
        city_for_excel = city or (self.config.target_region or '全国')
        row_data = (
            title,
            keyword,
            city_for_excel,
            contact_url,
            contact.get('member_name', ''),
            contact.get('mobile', ''),
            contact.get('address', ''),
        )
        for col, val in enumerate(row_data, start=1):
            self.ws.cell(row=self.excel_row, column=col, value=val)
        self.excel_row += 1
        self.wb.save(self.output_file)

    def _checkCollectLimit(self):
        """检查是否达到本次运行的采集上限。"""
        if self.config.total_max_shops > 0 and self.total_collected >= self.config.total_max_shops:
            print(f'已达采集上限 {self.config.total_max_shops}，停止后续采集')
            self._stop = True

    def _maybeRecoveryWait(self) -> None:
        """反爬：若处于访问被拒/验证码后的恢复期，额外等待再继续。"""
        if time.time() < self._recovery_until:
            extra = self._recovery_until - time.time()
            if extra > 0:
                print(f'恢复期等待 {extra:.0f} 秒...')
                time.sleep(extra)
            self._recovery_until = 0

    def _throttle(self, shop_start: float):
        """控制单条采集最少耗时，降低触发风控的概率。"""
        elapsed = time.time() - shop_start
        target = MIN_SECONDS_PER_SHOP + random.uniform(
            -MIN_SECONDS_PER_SHOP_JITTER, MIN_SECONDS_PER_SHOP_JITTER
        )
        target = max(target, MIN_SECONDS_PER_SHOP * 0.8)
        extra = max(0, target - elapsed)
        if extra > 0:
            time.sleep(extra)

    # ── 翻页 ─────────────────────────────────────────────

    def _goToNextPage(self, keyword: str, city: str, page_num: int) -> bool:
        """尝试点击下一页，返回 True 表示成功翻页，False 表示无下一页。"""
        city_desc = city or '整省/全国'
        page_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a.fui-next')
        if not page_elements:
            print(f'关键词「{keyword}」城市「{city_desc}」没有下一页，结束该任务')
            return False

        next_btn = page_elements[0]
        try:
            cls = (next_btn.get_attribute('class') or '').lower()
            disabled = (next_btn.get_attribute('aria-disabled') or '').lower()
            if 'disabled' in cls or 'fui-next-disabled' in cls or disabled == 'true':
                print(f'关键词「{keyword}」城市「{city_desc}」已是最后一页')
                return False
        except Exception:
            pass

        randomDelay(PAGE_TURN_WAIT_MIN, PAGE_TURN_WAIT_MAX)
        self._scrollToBottom()
        randomDelay(PAGE_OPEN_DELAY_MAX, PAGE_OPEN_DELAY_MAX + 1)
        self._closeKnownPopups()
        self.driver.execute_script('arguments[0].click();', next_btn)
        randomDelay(PAGE_TURN_WAIT_MIN, PAGE_TURN_WAIT_MAX)
        return True

    # ── DOM 联系信息提取 ──────────────────────────────────

    def _extractContactByDom(self) -> tuple[str, str, str, str, str]:
        """
        从当前页面 DOM 中提取联系人、电话、手机、传真、地址。
        优先匹配 1688 联系块结构（data-spm-anchor-id 容器），依次退化到 dl/dt/dd、class 关键词、全文正则。
        返回 (联系人, 电话, 手机, 传真, 地址)。
        """
        member_name = ''
        tel = ''
        mobile = ''
        fax = ''
        address = ''

        try:
            # 方法 0：1688 联系块结构 data-spm-anchor-id 容器
            tel, mobile, fax, address = self._extractFromSpmContainer()

            # 方法 1：dl > dt + dd 配对
            if not any([tel, mobile, fax, address]):
                member_name, tel, mobile, fax, address = self._extractFromDlDtDd()

            # 方法 2：class 含关键词的块
            if not member_name or not tel or not address:
                mn, tl, addr = self._extractFromClassKeywords()
                member_name = member_name or mn
                tel = tel or tl
                address = address or addr

            # 方法 3：全文文本正则
            if not all([member_name, tel, mobile, fax, address]):
                body_text = self.driver.find_element(By.TAG_NAME, 'body').text or ''
                r_name, r_tel, r_mobile, r_fax, r_addr = extractContactByRegex(body_text)
                member_name = member_name or r_name
                tel = tel or r_tel
                mobile = mobile or r_mobile
                fax = fax or r_fax
                address = address or r_addr
        except Exception:
            pass

        return (member_name or '', tel or '', mobile or '', fax or '', address or '')

    def _extractFromSpmContainer(self) -> tuple[str, str, str, str]:
        """从 data-spm-anchor-id 容器中提取电话、手机、传真、地址。"""
        tel = mobile = fax = address = ''
        try:
            containers = self.driver.find_elements(
                By.XPATH,
                f"//div[contains(@data-spm-anchor-id,'{SPM_ANCHOR_PREFIX}') "
                f"and .//div[contains(text(),'电话：')]]"
            )
            for container in containers:
                rows = container.find_elements(By.XPATH, './div')
                for row in rows:
                    parts = row.find_elements(By.XPATH, './div')
                    if len(parts) < 2:
                        continue
                    label = (parts[0].text or '').strip()
                    value_el = parts[1]
                    value = (value_el.text or '').strip()

                    if LABEL_ADDRESS in label or label in ('地址：', '地址'):
                        title_addr = value_el.get_attribute('title')
                        if title_addr and title_addr.strip():
                            value = title_addr.strip()
                        if value and not address:
                            address = value[:MAX_ADDRESS_LEN]
                    elif LABEL_TEL in label or label in ('电话：', '电话'):
                        if value and not tel:
                            tel = self._cleanPhone(value)
                    elif LABEL_MOBILE in label or label in ('手机：', '手机'):
                        if value and not mobile:
                            mobile = self._cleanPhone(value)
                    elif LABEL_FAX in label or label in ('传真：', '传真'):
                        if value and not fax:
                            fax = self._cleanPhone(value)

                if any([tel, mobile, fax, address]):
                    break
        except Exception:
            pass
        return tel, mobile, fax, address

    def _extractFromDlDtDd(self) -> tuple[str, str, str, str, str]:
        """从 dl > dt + dd 配对中提取联系信息。"""
        member_name = tel = mobile = fax = address = ''
        try:
            dts = self.driver.find_elements(By.TAG_NAME, 'dt')
            for dt in dts:
                label = (dt.text or '').strip()
                try:
                    dd = dt.find_element(By.XPATH, './following-sibling::dd[1]')
                    value = (dd.text or '').strip()
                except Exception:
                    value = ''
                if not value:
                    continue

                if LABEL_CONTACT in label or label == '联系人':
                    member_name = value[:30]
                elif LABEL_TEL in label or label == '电话':
                    digits = self._cleanPhone(value)
                    if len(re.sub(r'[^\d]', '', digits)) >= 5:
                        tel = digits
                elif LABEL_MOBILE in label or label == '手机':
                    digits = self._cleanPhone(value)
                    if len(re.sub(r'[^\d]', '', digits)) >= 5:
                        mobile = digits
                elif LABEL_FAX in label or label == '传真':
                    digits = self._cleanPhone(value)
                    if digits:
                        fax = digits
                elif LABEL_ADDRESS in label or label == '地址':
                    address = value[:MAX_ADDRESS_LEN]
        except Exception:
            pass
        return member_name, tel, mobile, fax, address

    def _extractFromClassKeywords(self) -> tuple[str, str, str]:
        """通过 class 含关键词的块提取联系人、电话、地址。"""
        member_name = tel = address = ''
        field_map = [
            ("//*[contains(@class,'member') or contains(@class,'contact-name')]", 'member'),
            ("//*[contains(@class,'tel') or contains(@class,'phone')]", 'tel'),
            ("//*[contains(@class,'address') or contains(@class,'addr')]", 'address'),
        ]
        for xpath, key in field_map:
            try:
                for el in self.driver.find_elements(By.XPATH, xpath):
                    t = (el.text or '').strip()
                    if not t or len(t) > 300:
                        continue
                    if key == 'member' and not member_name and re.match(r'^[\u4e00-\u9fa5a-zA-Z\s]{2,20}$', t):
                        member_name = t[:30]
                        break
                    if key == 'tel' and not tel and re.search(r'\d{5,}', t):
                        tel = re.sub(r'\s+', '-', re.sub(r'[^\d\-]', '', t))[:50] or t[:50]
                        break
                    if key == 'address' and not address:
                        address = t[:MAX_ADDRESS_LEN]
                        break
            except Exception:
                pass
        return member_name, tel, address

    @staticmethod
    def _cleanPhone(raw: str) -> str:
        """清洗电话号码字符串，仅保留数字和横线。"""
        return re.sub(r'\s+', ' ', re.sub(r'[^\d\-\s]', '', raw)).strip()[:50]

    # ── 弹窗处理 ─────────────────────────────────────────

    def _closeKnownPopups(self):
        """关闭已知冗余弹窗（baxia 遮罩/对话框），不关闭滑块验证框。"""
        try:
            self.driver.execute_script(
                "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
            )
            self.driver.execute_script(JS_REMOVE_BAXIA_MASK)
            self.driver.execute_script(JS_REMOVE_BAXIA_DIALOG)
        except Exception:
            pass

    def _closeAccessDeniedPopup(self) -> bool:
        """
        检测并尝试关闭「亲，访问被拒绝」弹窗。
        关闭后进行冷却等待并刷新页面。返回 True 表示曾检测到该弹窗。
        """
        try:
            body_text = self.driver.find_element(By.TAG_NAME, 'body').text or ''
            if TEXT_ACCESS_DENIED not in body_text:
                return False

            print('检测到「访问被拒绝」弹窗，进入冷却等待...')
            closed = False
            for _ in range(MAX_ACCESS_DENIED_CLOSE_ATTEMPTS):
                try:
                    deny_els = self.driver.find_elements(
                        By.XPATH, f"//*[contains(text(),'{TEXT_ACCESS_DENIED}')]"
                    )
                    for el in deny_els:
                        try:
                            parent = el.find_element(
                                By.XPATH,
                                "./ancestor::*[contains(@class,'dialog') or "
                                "contains(@class,'modal') or contains(@class,'popup')][1]"
                            )
                            close_btns = parent.find_elements(
                                By.XPATH,
                                ".//*[contains(@class,'close') or text()='×' or "
                                "text()='关闭' or contains(text(),'×')]"
                            )
                            if close_btns:
                                self.driver.execute_script('arguments[0].click();', close_btns[0])
                                closed = True
                                time.sleep(1)
                                break
                        except Exception:
                            pass
                    if closed:
                        break

                    script = f"""
                    var text = '{TEXT_ACCESS_DENIED}';
                    var all = document.querySelectorAll('div, section');
                    for (var i = all.length - 1; i >= 0; i--) {{
                        var el = all[i];
                        if (el.innerText && el.innerText.indexOf(text) !== -1) {{
                            var p = el.closest('.dialog') || el.closest('.modal')
                                || el.closest('[class*="dialog"]') || el.closest('[class*="modal"]')
                                || el.parentElement;
                            if (p) {{ p.remove(); return true; }}
                        }}
                    }}
                    return false;
                    """
                    if self.driver.execute_script(script):
                        closed = True
                        break
                except Exception:
                    pass
                time.sleep(0.5)

            cooldown = random.uniform(ACCESS_DENIED_COOLDOWN_MIN, ACCESS_DENIED_COOLDOWN_MAX)
            print(f'冷却等待 {cooldown:.0f} 秒后刷新...')
            time.sleep(cooldown)
            try:
                self.driver.refresh()
                randomDelay(PAGE_OPEN_DELAY_MIN + 2, PAGE_OPEN_DELAY_MAX + 3)
            except Exception:
                pass
            # 反爬：进入恢复期，后续几次操作延长等待
            self._recovery_until = time.time() + random.uniform(
                RECOVERY_AFTER_DENIED_EXTRA_MIN, RECOVERY_AFTER_DENIED_EXTRA_MAX
            )
            return True
        except Exception:
            return False

    # ── 滑块验证码处理 ───────────────────────────────────

    def _waitCaptchaResolved(self) -> tuple[bool, int]:
        """
        检测并尝试自动解决滑块验证码，失败后等待手动操作。
        返回 (是否通过, 等待秒数)。
        """
        start = time.time()
        if not self._detectCaptcha():
            return True, 0

        print('检测到滑块验证码，尝试自动拖拽...')
        self._closeKnownPopups()

        # 尝试自动拖拽
        try:
            if self._tryAutoSolveCaptcha() and not self._detectCaptcha():
                return True, int(time.time() - start)
        except Exception as e:
            print(f'自动拖拽出错: {e}')
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

        # 回退到手动操作
        if self._detectCaptcha():
            print('自动拖拽未通过，请在浏览器中手动完成验证...')
            wait_captcha = WebDriverWait(
                self.driver, self.config.captcha_timeout, poll_frequency=3.0
            )

            def _gone(d):
                if self._detectCaptcha():
                    self._closeKnownPopups()
                    return False
                return True

            try:
                wait_captcha.until(_gone)
                return True, int(time.time() - start)
            except Exception:
                return False, int(time.time() - start)

        return True, int(time.time() - start)

    def _detectCaptcha(self) -> bool:
        """检测当前页面是否存在滑块验证码。"""
        try:
            body_text = self.driver.find_element(By.TAG_NAME, 'body').text or ''
        except Exception:
            body_text = ''

        captcha_keywords = ['slide to verify', '滑动验证', '拖动', '完成验证']
        if any(kw in body_text.lower() for kw in captcha_keywords):
            return True

        try:
            for el in self.driver.find_elements(By.CSS_SELECTOR, '.baxia-dialog'):
                text = (el.text or '').strip()
                if text and ('拖动' in text or '验证' in text or 'slide' in text.lower()):
                    return True
        except Exception:
            pass
        return False

    def _tryAutoSolveCaptcha(self) -> bool:
        """尝试自动拖拽滑块，最多重试 SLIDER_MAX_RETRY 次。"""
        for attempt in range(SLIDER_MAX_RETRY):
            try:
                self.driver.switch_to.default_content()
            except Exception:
                pass

            slider = self._findBySelectors(SLIDER_BTN_SELECTORS)
            track = self._findBySelectors(SLIDER_TRACK_SELECTORS)

            # 主页面没找到，尝试切入 iframe
            in_iframe = False
            if not slider:
                if self._switchToSliderIframe():
                    in_iframe = True
                    slider = self._findBySelectors(SLIDER_BTN_SELECTORS)
                    track = self._findBySelectors(SLIDER_TRACK_SELECTORS)

            if not slider:
                time.sleep(1)
                continue

            distance = self._calcSliderDistance(slider, track)
            jitter = random.randint(-10, 10)
            distance = max(100, distance + jitter)
            logVerbose(f'自动拖拽滑块（第 {attempt + 1} 次，距离 {distance}px）')

            try:
                self._humanLikeDrag(slider, distance)
            except Exception as e:
                print(f'拖拽操作异常: {e}')

            if in_iframe:
                try:
                    self.driver.switch_to.default_content()
                except Exception:
                    pass

            randomDelay(1.5, 3.0)

            if not self._detectCaptcha():
                print('自动拖拽验证通过！')
                return True

            randomDelay(2.0, 4.0)
            self._clickSliderRefresh()

        return False

    def _calcSliderDistance(self, slider, track) -> int:
        """计算滑块拖拽距离：优先用轨道宽度，否则用默认值。"""
        if track:
            dist = track.size.get('width', SLIDER_DEFAULT_DISTANCE) - slider.size.get('width', 40)
            return max(dist, 100)
        return SLIDER_DEFAULT_DISTANCE

    def _clickSliderRefresh(self):
        """尝试点击滑块刷新/重置按钮。"""
        try:
            btn = self._findBySelectors([
                '.nc-lang-cnt .errloading a', '.errloading a',
                '#nc_1_refresh1', '.btn_reload',
            ])
            if btn:
                btn.click()
                randomDelay(1.0, 2.0)
        except Exception:
            pass

    def _switchToSliderIframe(self) -> bool:
        """尝试切换到滑块验证码所在的 iframe。"""
        for sel in SLIDER_IFRAME_SELECTORS:
            try:
                for iframe in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if iframe.is_displayed():
                        self.driver.switch_to.frame(iframe)
                        return True
            except Exception:
                continue

        # 通用兜底：遍历所有 iframe
        try:
            for iframe in self.driver.find_elements(By.TAG_NAME, 'iframe'):
                try:
                    if not iframe.is_displayed():
                        continue
                    self.driver.switch_to.frame(iframe)
                    if self._findBySelectors(SLIDER_BTN_SELECTORS):
                        return True
                    self.driver.switch_to.default_content()
                except Exception:
                    try:
                        self.driver.switch_to.default_content()
                    except Exception:
                        pass
        except Exception:
            pass
        return False

    def _humanLikeDrag(self, slider_el, distance: int):
        """模拟人类拖拽滑块：先加速后减速，带随机抖动。"""
        action = ActionChains(self.driver)
        action.move_to_element(slider_el)
        action.pause(random.uniform(0.3, 0.6))
        action.click_and_hold(slider_el)
        action.pause(random.uniform(0.1, 0.25))

        for dx, dy, dt_ms in self._generateHumanTrack(distance):
            action.move_by_offset(dx, dy)
            action.pause(dt_ms / 1000.0)

        action.pause(random.uniform(0.3, 0.8))
        action.release()
        action.perform()

    @staticmethod
    def _generateHumanTrack(distance: int) -> list[tuple[int, int, int]]:
        """生成模拟人类拖拽的轨迹点列表（先加速后减速，带随机抖动）。"""
        tracks: list[tuple[int, int, int]] = []
        current = 0
        accel_end = distance * 0.7

        # 加速段
        t = 0.0
        while current < accel_end:
            t += random.uniform(0.02, 0.04)
            progress = min(t / 0.5, 1.0)
            step = max(1, int(random.uniform(2, 6) * (1 + progress * 2)))
            if current + step > accel_end:
                step = max(1, int(accel_end - current))
            dy = random.choice([-1, 0, 0, 0, 1])
            dt = random.randint(SLIDER_STEP_MIN_MS, SLIDER_STEP_MAX_MS)
            tracks.append((step, dy, dt))
            current += step

        # 减速段
        remaining = distance - current
        while remaining > 0:
            step = max(1, int(remaining * random.uniform(0.15, 0.4)))
            if step > remaining:
                step = max(1, int(remaining))
            dy = random.choice([0, 0, 0, -1, 1])
            dt = random.randint(SLIDER_STEP_MAX_MS, SLIDER_STEP_MAX_MS * 3)
            tracks.append((step, dy, dt))
            remaining -= step

        return tracks

    # ── 页面滚动 ─────────────────────────────────────────

    def _scrollToBottom(self):
        """将页面滚动到底部。"""
        try:
            self.driver.execute_script(
                "window.scrollTo({top: document.body.scrollHeight, behavior: 'auto'});"
            )
        except Exception:
            try:
                self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
            except Exception:
                pass

    def _scrollToLoadAllResults(self):
        """渐进式滚动页面，触发 1688 搜索结果的懒加载。"""
        try:
            self.driver.execute_script('window.scrollTo(0, 0);')
            time.sleep(0.3)

            total_height = self.driver.execute_script('return document.body.scrollHeight;')
            current_pos = 0

            while current_pos < total_height:
                step = SCROLL_STEP_PX + random.randint(
                    -SCROLL_STEP_PX_JITTER, SCROLL_STEP_PX_JITTER
                )
                step = max(200, step)
                current_pos += step
                self.driver.execute_script(f'window.scrollTo(0, {current_pos});')
                randomDelay(SCROLL_STEP_WAIT_MIN, SCROLL_STEP_WAIT_MAX)

                try:
                    loaded = len(self.driver.find_elements(By.CSS_SELECTOR, 'a.company-name'))
                    if loaded >= EXPECTED_ITEMS_PER_PAGE:
                        break
                except Exception:
                    pass

                try:
                    total_height = self.driver.execute_script('return document.body.scrollHeight;')
                except Exception:
                    pass

            self._scrollToBottom()
            time.sleep(SCROLL_FINAL_WAIT)
            self.driver.execute_script('window.scrollTo(0, 0);')
            time.sleep(0.3)
        except Exception:
            pass

    def _scrollToElement(self, element):
        """将指定元素滚动到视口中间，便于观察采集进度。"""
        try:
            self.driver.execute_script(
                "arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});",
                element,
            )
        except Exception:
            pass

    # ── 工具方法 ─────────────────────────────────────────

    def _findBySelectors(self, selectors: list[str]):
        """按优先级依次尝试 CSS 选择器列表，返回第一个可见元素或 None。"""
        for sel in selectors:
            try:
                for el in self.driver.find_elements(By.CSS_SELECTOR, sel):
                    if el.is_displayed() and el.size.get('width', 0) > 0:
                        return el
            except Exception:
                continue
        return None

    def _closeTabAndReturn(self, main_window: str):
        """关闭当前标签并切回主窗口。"""
        try:
            randomDelay(2, 3)
            self.driver.close()
            self.driver.switch_to.window(main_window)
        except Exception:
            pass

    def _safeCloseAndReturn(self, main_window: str):
        """安全关闭非主窗口标签并切回主窗口（异常恢复用）。"""
        try:
            randomDelay(2, 3)
            if self.driver.current_window_handle != main_window:
                self.driver.close()
            self.driver.switch_to.window(main_window)
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════
#  入口
# ═══════════════════════════════════════════════════════════

def main():
    """脚本主入口：加载配置 → 用户交互 → 登录 → 采集。"""
    config = loadRuntimeConfig()
    config = collectUserInput(config)

    with AlibabaScraper(config) as scraper:
        scraper.login()
        scraper.run()


if __name__ == '__main__':
    freeze_support()
    main()
